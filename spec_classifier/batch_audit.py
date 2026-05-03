"""
batch_audit.py v2 — Batch pipeline validation + LLM prediction layer.

Finds all *_annotated.xlsx in OUTPUT folder, runs:
  1. Rule-based checks (E1–E18) — fast, local
  2. LLM prediction — Claude looks at Option Name and predicts entity_type + device_type.
     If prediction disagrees with pipeline → adds AI_MISMATCH pометку.

Usage:
    python batch_audit.py --output-dir C:/Users/G/Desktop/OUTPUT
    python batch_audit.py --output-dir C:/Users/G/Desktop/OUTPUT --no-ai     # rule checks only
    python batch_audit.py --output-dir C:/Users/G/Desktop/OUTPUT --vendor hpe
    python batch_audit.py --output-dir C:/Users/G/Desktop/OUTPUT --since 2026-03-04
    python batch_audit.py --output-dir C:/Users/G/Desktop/OUTPUT --dry-run

Requirements:
    pip install anthropic openpyxl pandas
"""

import argparse
import json
import os
import re
import sys
import time
from datetime import datetime
from pathlib import Path

import yaml
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

if hasattr(sys.stdout, "reconfigure") and sys.stdout.encoding \
        and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure") and sys.stderr.encoding \
        and sys.stderr.encoding.lower() not in ("utf-8", "utf8"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Constants ─────────────────────────────────────────────────────────────────

HW_TYPE_VOCAB = frozenset({
    # Cycle 2 (PR-8–PR-10): no new hw_type members — only device_type refinements
    # mapped through vendor YAML device_type_map + batch_audit DEVICE_TYPE_ALIASES.
    "server", "switch", "storage_system", "wireless_ap",
    "cpu", "memory", "gpu",
    "storage_drive", "storage_enclosure", "storage_controller", "hba", "backplane", "io_module",
    "network_adapter", "transceiver", "cable",
    "psu", "fan", "heatsink",
    "riser", "chassis", "rail", "blank_filler",
    "management", "tpm", "accessory",
})

VALID_ENTITY_TYPES = {"BASE", "HW", "CONFIG", "SOFTWARE", "SERVICE", "LOGISTIC", "NOTE", "UNKNOWN"}
VALID_STATES = {"PRESENT", "ABSENT", "DISABLED", ""}


def _load_config() -> dict:
    """Load config.yaml, overlay with config.local.yaml (paths only)."""
    base = Path(__file__).resolve().parent / "config.yaml"
    with open(base, encoding="utf-8-sig") as f:
        cfg = yaml.safe_load(f) or {}
    local = base.parent / "config.local.yaml"
    if local.exists():
        with open(local, encoding="utf-8-sig") as f:
            local_cfg = yaml.safe_load(f) or {}
        cfg.update(local_cfg)
    return cfg


def _get_known_vendors(config: dict) -> list[str]:
    """Return sorted list of vendor keys from config."""
    return sorted(config.get("vendor_rules", {}).keys())


def _load_device_type_maps(config: dict) -> dict[str, dict]:
    """Load device_type→hw_type maps from vendor YAML rules."""
    maps = {}
    for vendor, rules_path in config.get("vendor_rules", {}).items():
        path = Path(__file__).resolve().parent / rules_path
        if path.exists():
            with open(path, encoding="utf-8") as f:
                data = yaml.safe_load(f)
            htr = data.get("hw_type_rules") or {}
            maps[vendor] = htr.get("device_type_map") or {}
    return maps


try:
    _device_type_map: dict[str, dict] = _load_device_type_maps(_load_config())
except Exception:
    _device_type_map: dict[str, dict] = {}

# ── E4 state validation config ────────────────────────────────────────────────

E4_STATE_VALIDATORS: dict = {
    "dell": {
        "type": "pattern_checks",
        "rules": [
            {
                "pattern": r"^\s*No\s+",
                "expected_states": ["ABSENT"],
                "override_pattern": r"\b\d+\s+Rear\s+Blanks?\b",
            },
            {"pattern": r"\bDisabled\b", "expected_states": ["DISABLED", "ABSENT"]},
            {"pattern": r"\b(None|Empty|Without)\b", "expected_states": ["ABSENT"]},
        ],
    },
    "cisco": {
        "type": "pattern_checks",
        "rules": [
            {"pattern": r"(?i)No .+ Selected", "expected_states": ["ABSENT"]},
        ],
    },
    "hpe": {
        "type": "unexpected_states",
        "states": ["ABSENT", "DISABLED"],
    },
    "lenovo": {
        "type": "unexpected_states",
        "states": ["ABSENT", "DISABLED"],
    },
    "huawei": {
        "type": "unexpected_states",
        "states": ["ABSENT", "DISABLED"],
    },
    "xfusion": {
        "type": "unexpected_states",
        "states": ["ABSENT", "DISABLED"],
    },
}


def _check_e4(vendor: str, option_name: str, state: str, issues: list[str]) -> None:
    import re
    cfg = E4_STATE_VALIDATORS.get(vendor)
    if not cfg:
        return
    if cfg["type"] == "unexpected_states":
        if state in cfg["states"]:
            issues.append(f"E4:state_unexpected_for_{vendor}[{state}]")
    elif cfg["type"] == "pattern_checks":
        for rule in cfg["rules"]:
            if re.search(rule["pattern"], option_name, re.IGNORECASE):
                override = rule.get("override_pattern")
                if override and re.search(override, option_name, re.IGNORECASE):
                    continue
                if state not in rule["expected_states"]:
                    issues.append(
                        f"E4:state_mismatch[expected:{rule['expected_states'][0]},got:{state}]"
                    )
                break


# Colours
RED    = "FFC7CE"
ORANGE = "FFE0B2"
YELLOW = "FFFDE7"
PURPLE = "E1BEE7"
GREEN  = "C6EFCE"
TEAL   = "B2EBF2"   # AI mismatch
WHITE  = "FFFFFF"

# ── LLM prediction layer ──────────────────────────────────────────────────────

def _build_llm_system(known_vendors: list[str]) -> str:
    vendor_list = ", ".join(v.upper() if len(v) <= 4 else v.title() for v in known_vendors)
    return f"""You are a hardware spec classifier for enterprise IT equipment ({vendor_list}).
Given a list of rows with Option Name (and optionally Module Name), predict for each row:""" \
        + _LLM_SYSTEM_BODY


_LLM_SYSTEM_BODY = """
- entity_type: one of BASE, HW, CONFIG, SOFTWARE, SERVICE, LOGISTIC, NOTE
- device_type: one of cpu, memory, gpu, storage_nvme, storage_ssd, storage_hdd,
  storage_controller, storage_enclosure, hba, network_adapter, transceiver, cable, sfp_cable, fiber_cable,
  psu, fan, heatsink, riser, chassis, rail, blank_filler, management, tpm, accessory,
  power_cord, raid_controller, nic, ram, drive_cage, backplane, motherboard, bezel, battery,
  front_panel, power_distribution_board, interconnect_board, io_module, media_bay,
  air_duct, optical_drive,
  server, switch, storage_system, wireless_ap
  OR empty string "" if not applicable.

front_panel = server front operator panel with LEDs/LCD/status display, including
diagnostic panels and Operator Panel ASM modules. Maps to hw_type=management.
tpm covers BOTH TPM 2.0 modules AND Root of Trust (RoT) security modules — Lenovo
documents these together as "Firmware and Root of Trust/TPM 2.0 Security Module".
power_distribution_board = internal server PDB / Power Interface / Power Interconnect
Board that distributes power within the chassis. Distinct from PSU (which converts AC
to DC) and from power_cord. Maps to hw_type=chassis.
interconnect_board = internal server PCIe Switch / Retimer / I/O Board (e.g. "PCIe
Switch Board with two 144-lanes Switches", "PCIe Retimer Card", "System I/O Board").
Distinct from motherboard (main system board) and from io_module (external pluggable
storage I/O module). Maps to hw_type=chassis.
media_bay = server media bay — a removable physical bay that can host front I/O panel,
optical drive, or other front-mounted devices (e.g. "Standard Media Bay", "Media Bay
without External Diagnostics Port"). Distinct from drive_cage (which is drive-specific
and maps to backplane). Maps to hw_type=chassis.
air_duct = server air duct / airflow guide — plastic or sheet-metal directional
component that channels airflow over CPUs, DIMMs, GPUs (e.g. "ThinkSystem 2U Main
Air Duct", "Air duct (1U radiator)"). Distinct from heatsink (heat-conducting metal
fin stack) and from fan (active cooling). Maps to hw_type=accessory (no cooling
hw_type exists in the canonical 26-value vocab).
optical_drive = optical disk drive — DVD-RW / DVD-ROM / CD-ROM / Blu-ray / ODD,
removable storage media reader (e.g. "HPE 9.5mm SATA DVD-RW Optical Drive"). Maps
to hw_type=storage_drive (same bucket as HDD/SSD/NVMe storage drives).

Rules:
- BASE = the main product itself (server model, switch model, storage array)
- HW = physical hardware component (cpu, memory, drive, nic, psu, fan, cable, etc.)
- CONFIG = factory integration flag, enablement kit, configuration tracking, jumper, dummy PID for selection
- SOFTWARE = OS, firmware, license, software kit
- SERVICE = warranty, support, carepack, deployment service
- LOGISTIC = shipping labels, delivery/freight charges, packing materials, documentation only
- NOTE = informational text, no physical item

BASE is determined by row position and module context (bundle root, first product row). If pipeline says BASE, trust it — do not reclassify based on option_name alone.
For Dell BOM rows, Module Name is the PRIMARY signal for entity_type. 'Chassis Configuration' module contains HW rows, not CONFIG.
Power cords, rack rails, mounting hardware, and cables are ALWAYS HW (entity_type=HW), never LOGISTIC. LOGISTIC is ONLY for non-physical items: shipping charges, freight, labels, packing slips.
For CONFIG: "Dummy PID", "Airflow Selection", "Enablement Kit", "Factory Integrated", "FIO" → entity_type=CONFIG.
For blank fillers and bezel blanks → entity_type=HW, device_type=blank_filler.
For rows starting with 'No ...' (e.g., 'No Hard Drive', 'No TPM', 'No Quick Sync', 'No Rear Storage'): entity_type=HW, state=ABSENT. These represent absent hardware slots, NOT informational notes.

Respond ONLY with a JSON array, one object per input row, in same order:
[{"row_index": 0, "predicted_entity": "HW", "predicted_device_type": "cpu", "confidence": "high"},...]
confidence: "high" if obvious, "medium" if plausible, "low" if uncertain.
No markdown, no explanation, just the JSON array."""

LLM_SYSTEM = _build_llm_system(_get_known_vendors(_load_config()))


# GPT-4o-mini pricing per 1M tokens (update if model changes)
PRICING = {
    "gpt-4o-mini":        {"in": 0.150, "out": 0.600},
    "gpt-4o":             {"in": 2.50,  "out": 10.00},
    "claude-opus-4-5":    {"in": 15.00, "out": 75.00},
    "claude-sonnet-4-5":  {"in": 3.00,  "out": 15.00},
}

def predict_batch(rows: list[dict], client, llm_model: str = "gpt-4o-mini") -> tuple[list, int, int]:
    """Send batch of rows to prediction. Returns (results, tokens_in, tokens_out)."""
    if not rows:
        return [], 0, 0

    payload = [
        {
            "row_index": i,
            "option_name": r.get("option_name") or r.get("Option Name") or "",
            "module_name": r.get("module_name") or r.get("Module Name") or "",
        }
        for i, r in enumerate(rows)
    ]

    try:
        if type(client).__name__ == "OpenAI":
            response = client.chat.completions.create(
                model=llm_model,
                max_tokens=4096,
                messages=[
                    {"role": "system", "content": LLM_SYSTEM},
                    {"role": "user", "content": json.dumps(payload, ensure_ascii=False)},
                ],
            )
            text = response.choices[0].message.content.strip()
            tok_in = response.usage.prompt_tokens if response.usage else 0
            tok_out = response.usage.completion_tokens if response.usage else 0
        else:
            response = client.messages.create(
                model=llm_model,
                max_tokens=4096,
                system=LLM_SYSTEM,
                messages=[{"role": "user", "content": json.dumps(payload, ensure_ascii=False)}],
            )
            text = response.content[0].text.strip()
            tok_in = response.usage.input_tokens if response.usage else 0
            tok_out = response.usage.output_tokens if response.usage else 0
        # Strip markdown fences if present
        if text.startswith("```"):
            text = text.split("```")[1]
            if text.startswith("json"):
                text = text[4:]
        parsed = json.loads(text.strip())
        if isinstance(parsed, dict):
            for v in parsed.values():
                if isinstance(v, list):
                    return v, tok_in, tok_out
            return [], tok_in, tok_out
        return parsed, tok_in, tok_out
    except Exception as e:
        print(f"    ⚠ LLM batch error: {e}")
        return [], 0, 0


def run_llm_predictions(data_rows: list[dict], client, batch_size: int = 40, llm_model: str = "gpt-4o-mini") -> tuple[dict, int, int]:
    """
    Run LLM predictions on all ITEM rows in batches.
    Returns dict: original_df_index -> prediction.
    """
    results = {}
    total_tok_in = 0
    total_tok_out = 0
    # Predict only on component rows: HW, BASE, UNKNOWN
    # Skip SERVICE/LOGISTIC/NOTE/CONFIG/SOFTWARE — rule-checks sufficient
    # Also skip: header rows, rows without option_name, Factory Integrated rows
    AI_WORTHY = {"HW", "BASE", "UNKNOWN", ""}
    to_predict = [
        (i, r) for i, r in enumerate(data_rows)
        if str(r.get("row_kind") or "ITEM").upper() != "HEADER"
        and (r.get("option_name") or r.get("Option Name") or "").strip()
        and str(r.get("entity_type") or r.get("Entity Type") or "").strip().upper() in AI_WORTHY
        and not str(r.get("option_name") or "").strip().lower().startswith("factory integrated")
        and str(r.get("is_factory_integrated") or r.get("is_factor") or "").strip().lower() != "true"
    ]
    skipped = len(data_rows) - len(to_predict)
    skip_note = f"  (пропущено {skipped} строк)" if skipped else ""

    total_batches = (len(to_predict) + batch_size - 1) // batch_size
    bar_width = 30

    def print_bar(done: int, total: int, extra: str = ""):
        pct = done / total if total else 1
        filled = int(bar_width * pct)
        bar = "█" * filled + "░" * (bar_width - filled)
        print(f"\r    🤖 [{bar}] {done}/{total} батчей {extra}   ", end="", flush=True)

    print(f"    🤖 LLM: {len(to_predict)} строк → {total_batches} батчей{skip_note}", flush=True)
    print_bar(0, total_batches)

    for batch_idx, batch_num in enumerate(range(0, len(to_predict), batch_size)):
        chunk = to_predict[batch_num:batch_num + batch_size]
        index_map = {j: orig_i for j, (orig_i, _) in enumerate(chunk)}
        batch_rows = [r for _, r in chunk]

        t0 = time.time()
        preds, tok_in, tok_out = predict_batch(batch_rows, client, llm_model)
        elapsed = time.time() - t0
        total_tok_in += tok_in
        total_tok_out += tok_out

        for pred in preds:
            local_idx = pred.get("row_index", -1)
            if local_idx in index_map:
                results[index_map[local_idx]] = pred

        done = batch_idx + 1
        remaining = (total_batches - done) * elapsed
        eta = f"ETA {remaining:.0f}s" if remaining > 1 else "почти готово"
        print_bar(done, total_batches, eta)

        if batch_num + batch_size < len(to_predict):
            time.sleep(0.3)

    print(f"\r    🤖 ✅ {len(results)}/{len(to_predict)} строк обработано" + " " * 20, flush=True)
    return results, total_tok_in, total_tok_out


# device_type aliases — semantic equivalences for AI_MISMATCH suppression ONLY.
# NOT an hw_type map. (e.g. drive_cage→backplane: AI saying "backplane" matches
# pipeline saying "drive_cage" without raising an AI_MISMATCH flag. Per-vendor
# hw_type maps live in rules/<vendor>_rules.yaml device_type_map.)
DEVICE_TYPE_ALIASES = {
    "ram":                       "memory",
    "nic":                       "network_adapter",
    "raid_controller":           "storage_controller",
    "sfp_cable":                 "cable",
    "fiber_cable":               "cable",
    "drive_cage":                "backplane",
    "bezel":                     "chassis",
    "motherboard":               "chassis",
    "storage_nvme":              "storage_drive",
    "storage_ssd":               "storage_drive",
    "storage_hdd":               "storage_drive",
    "power_cord":                "cable",
    "hba":                       "storage_controller",
    "front_panel":               "management",  # PR-9a Q7a: Front Operator Panel maps to management bucket
    "power_distribution_board":  "chassis",     # PR-9b Q8: PDB / Power Interface / Power Interconnect Board → chassis
    "interconnect_board":        "chassis",     # PR-9b Q8: PCIe Switch / Retimer / I/O Board → chassis
    "media_bay":                 "chassis",     # PR-9c Q9: Media Bay (front-panel/optical/etc bay) → chassis
    "air_duct":                  "accessory",   # PR-10 Q10e: Air Duct / Airduct (airflow guide) → accessory bucket
    "optical_drive":             "storage_drive",  # PR-10 Q10f: DVD-RW / Optical Drive → storage_drive bucket
}

# Entity типы где пайплайн всегда прав — AI не знает бизнес-логику
# "Disabled" строка это HW ABSENT а не NOTE; chassis это физика а не CONFIG
ENTITY_TRUST_PIPELINE = {"LOGISTIC", "SOFTWARE", "SERVICE", "CONFIG", "BASE"}
# hw_type значения где entity=HW всегда правильный (chassis — физический корпус)
HW_TYPE_TRUST = {"chassis", "backplane", "riser", "rail", "battery", "accessory", "blank_filler", "storage_enclosure"}

# device_type значения где пайплайн всегда прав — AI часто ошибается
# cable kit ≠ accessory, battery/capacitor ≠ accessory, rail ≠ accessory
# front_panel — узкий regex (Front Operator Panel / LCD system info display),
# AI часто кидает в accessory.
# power_distribution_board / interconnect_board (PR-9b Q8) — узкие regex
# по "Board" suffix; AI часто плывёт в accessory или motherboard.
# air_duct / optical_drive (PR-10 Q10e/Q10f) — узкие regex; AI часто кидает
# air_duct в accessory (теряется новый device_type) и optical_drive в accessory
# (вместо storage_drive bucket).
DEVICE_TYPE_TRUST = {"cable", "battery", "rail", "riser", "blank_filler",
                     "accessory", "chassis", "backplane", "storage_enclosure",
                     "motherboard", "bezel", "front_panel",
                     "power_distribution_board", "interconnect_board",
                     "media_bay", "drive_cage",
                     "air_duct", "optical_drive"}

def build_ai_mismatch(pipeline_entity: str, pipeline_device: str,
                       pred: dict) -> str | None:
    """
    4 цели:
    1. AI определяет тип устройства самостоятельно
    2. Расхождение с пайплайном → AI_MISMATCH (требует внимания)
    3. Пайплайн не определил device_type → AI_SUGGEST (AI предлагает вариант)
    4. AI не уверен или видит аномалию → MANUAL_CHECK
    """
    if not pred:
        return None

    p_entity = pred.get("predicted_entity", "").upper()
    p_device = pred.get("predicted_device_type", "").lower()
    confidence = pred.get("confidence", "low")

    tags = []

    # ── Цель 2: расхождение entity ────────────────────────────────────────────
    # Пропускаем: low confidence, trusted entity types, trusted hw_types
    if (p_entity
            and confidence != "low"
            and pipeline_entity not in ("UNKNOWN", "")
            and pipeline_entity.upper() not in ENTITY_TRUST_PIPELINE
            and pipeline_device.lower() not in HW_TYPE_TRUST
            and p_entity != pipeline_entity.upper()):
        conf_tag = f"[{confidence}]" if confidence != "high" else ""
        tags.append(f"AI_MISMATCH{conf_tag}:entity[pipeline:{pipeline_entity}→ai:{p_entity}]")

    # ── Цели 2 и 3: device_type ───────────────────────────────────────────────
    if p_device and confidence != "low":
        pipe_dev = pipeline_device.lower()
        pipe_norm = DEVICE_TYPE_ALIASES.get(pipe_dev, pipe_dev)
        ai_norm   = DEVICE_TYPE_ALIASES.get(p_device, p_device)

        if pipe_dev and pipe_norm != ai_norm:
            # Цель 2: пайплайн поставил одно, AI говорит другое
            if pipe_dev in DEVICE_TYPE_TRUST:
                # Пайплайн доверен для этих типов — AI часто ошибается на Kit/Capacitor/etc
                tags.append(f"AI_MISMATCH[medium]:device_type[pipeline:{pipe_dev}→ai:{p_device}]")
            else:
                conf_tag = f"[{confidence}]" if confidence != "high" else ""
                tags.append(f"AI_MISMATCH{conf_tag}:device_type[pipeline:{pipeline_device}→ai:{p_device}]")
        elif not pipe_dev and pipeline_entity in ("HW", "BASE"):
            # Цель 3: пайплайн не определил device_type, AI предлагает
            if confidence == "high":
                tags.append(f"AI_SUGGEST:device_type={p_device}")
            else:
                # Цель 4: AI неуверен — пометить для ручной проверки
                tags.append(f"MANUAL_CHECK:device_type_unknown[ai_guess:{p_device}]")

    return "; ".join(tags) if tags else None


# ── Rule-based checks (E1–E18) ────────────────────────────────────────────────

def validate_row(row: dict, vendor: str,
                  device_type_map: dict[str, dict] | None = None) -> list[str]:
    import re
    issues = []

    entity = str(row.get("entity_type") or row.get("Entity Type") or "").strip().upper()
    state = str(row.get("state") or row.get("State") or "").strip().upper()
    hw_type = str(row.get("hw_type") or "").strip().lower()
    device_type = str(row.get("device_type") or "").strip().lower()
    option_name = str(row.get("option_name") or row.get("Option Name") or "").strip()
    row_kind = str(row.get("row_kind") or "").strip().upper()

    if row_kind == "HEADER":
        return []

    # E1 — invalid entity type
    if entity and entity not in VALID_ENTITY_TYPES:
        issues.append(f"E1:invalid_entity[{entity}]")

    # E2 — UNKNOWN
    if entity == "UNKNOWN":
        issues.append("E2:UNKNOWN_no_rule")

    # E3 — invalid state
    if state and state not in VALID_STATES:
        issues.append(f"E3:invalid_state[{state}]")

    # E4 — state logic by vendor (data-driven)
    _check_e4(vendor, option_name, state, issues)

    # E5 — hw_type on non-HW
    if entity not in ("HW", "BASE") and hw_type:
        issues.append(f"E5:hw_type_on_non_hw[entity:{entity}]")

    # E6 — device_type on wrong entity
    if entity not in ("HW", "LOGISTIC", "BASE") and device_type:
        issues.append(f"E6:device_type_on_wrong_entity[entity:{entity}]")

    # E7 — hw_type not in vocab
    if hw_type and hw_type not in HW_TYPE_VOCAB:
        issues.append(f"E7:hw_type_not_in_vocab[{hw_type}]")

    # E8 — HW missing hw_type
    _E8_NO_HW_TYPE_DEVICES = {"power_cord", "enablement_kit"}
    if entity == "HW" and state == "PRESENT" and not hw_type \
            and device_type not in _E8_NO_HW_TYPE_DEVICES:
        issues.append("E8:hw_type_missing_on_hw")

    # E9 — device_type → hw_type mapping
    if device_type and entity == "HW":
        _dtm = device_type_map if device_type_map is not None else _device_type_map
        mapping = _dtm.get(vendor, {})
        if device_type in mapping:
            expected_hw = mapping[device_type]
            if not hw_type:
                issues.append(f"E9:hw_type_missing_for_device_type[device:{device_type},expected:{expected_hw}]")
            elif hw_type != expected_hw:
                issues.append(f"E9:hw_type_mapping_mismatch[device:{device_type},expected:{expected_hw},got:{hw_type}]")

    # E10 — hw/device_type on BASE
    if entity == "BASE":
        if hw_type:
            issues.append(f"E10:hw_type_on_base[{hw_type}]")
        # device_type on BASE is valid — set by BASE-*-DT-* YAML rules.
        # Removed device_type sub-check (was producing false positives).

    # E11 — hw_type on CONFIG
    if entity == "CONFIG" and hw_type:
        issues.append(f"E11:hw_type_on_config[{hw_type}]")

    # E12 — hw/device_type on NOTE
    if entity == "NOTE" and (hw_type or device_type):
        issues.append("E12:hw_type_or_device_type_on_note")

    # E13 — LOGISTIC with physical cable/power_cord (should be HW)
    if entity == "LOGISTIC" and device_type in ("power_cord", "cable", "sfp_cable", "fiber_cable"):
        issues.append(f"E13:logistic_with_physical_cable[device:{device_type}→consider_HW]")

    # E14 — CONFIG with Dummy/Blank in name but no device_type
    if entity == "CONFIG" and not device_type:
        # ИСКЛЮЧЕНИЕ: NXK-AF-PE — Dummy PID для airflow selection, намеренно CONFIG без device_type
        _E14_EXCLUDE_SKUS = {"NXK-AF-PE"}
        _e14_sku_raw = row.get("skus") or row.get("sku") or row.get("part_number") or ""
        _e14_skus = {_e14_sku_raw} if isinstance(_e14_sku_raw, str) else set(_e14_sku_raw or [])
        if re.search(r"\b(dummy|blank|filler)\b", option_name, re.IGNORECASE) and not _E14_EXCLUDE_SKUS.intersection(_e14_skus):
            issues.append("E14:config_looks_like_blank_filler[suggest:device_type=blank_filler]")

    # E15 — BASE without device_type (info, not error)
    if entity == "BASE" and not device_type:
        issues.append("E15:info_base_no_device_type")

    # E16 — blank_filler + ABSENT: слот занят заглушкой, диск отсутствует
    # Не ошибка — но важно видеть: заглушка физически есть, диска нет
    # ИСКЛЮЧЕНИЕ: 412-AASK (NIC-slot blank) и 470-BCHP (BOSS-slot blank) —
    # не drive bay fillers, E16 семантически неприменим. Узкое SKU-исключение.
    _E16_EXCLUDE_SKUS = {"412-AASK", "470-BCHP"}
    _sku_value = row.get("skus") or row.get("part_number") or row.get("sku")
    _row_skus = {_sku_value} if isinstance(_sku_value, str) else set(_sku_value or [])
    if device_type == "blank_filler" and state == "ABSENT" and not _E16_EXCLUDE_SKUS.intersection(_row_skus):
        issues.append("E16:info_slot_has_filler_no_drive")

    # E17 — HW строка без device_type и без hw_type (пайплайн не определил тип)
    # Отличается от E8 (E8 = hw_type пустой у PRESENT HW)
    if entity == "HW" and not device_type and not hw_type and state == "PRESENT":
        issues.append("E17:info_hw_type_not_determined")

    # E18: LOGISTIC row with physical keyword but no device_type
    PHYSICAL_KEYWORDS = {"cord", "cable", "rail", "bracket", "mount", "kit", "rack", "pdu", "ups"}
    if entity == "LOGISTIC":
        name_lower = (row.get("option_name") or "").lower()
        if any(kw in name_lower for kw in PHYSICAL_KEYWORDS) and not device_type:
            issues.append(f"E18:logistic_with_physical_keyword[no_device_type:{row.get('option_name')}]")

    return issues


def issue_color(issues_str: str) -> str:
    if not issues_str or issues_str == "OK":
        return GREEN
    if "E2:" in issues_str:
        return RED                  # красный — нет правила
    if "AI_MISMATCH" in issues_str:
        return TEAL                 # голубой — AI не согласен
    if "MANUAL_CHECK" in issues_str:
        return "FFE0B2"             # оранжевый — ручная проверка
    if "AI_SUGGEST" in issues_str:
        return "E8F5E9"             # светло-зелёный — AI предлагает вариант
    if any(c in issues_str for c in ["E5:", "E6:", "E10:", "E11:", "E12:"]):
        return PURPLE               # фиолетовый — логика entity
    if any(c in issues_str for c in ["E8:", "E9:", "E7:", "E13:", "E14:"]):
        return ORANGE               # оранжевый — hw_type проблемы
    if "E4:" in issues_str:
        return YELLOW               # жёлтый — state mismatch
    if "E16:" in issues_str:
        return "FFF9C4"             # светло-жёлтый — заглушка в слоте
    if "E17:" in issues_str:
        return "FAFAFA"             # светло-серый — тип не определён
    if "E18:" in issues_str:
        return YELLOW               # жёлтый — LOGISTIC с физическим keyword
    if "E15:" in issues_str:
        return "F5F5F5"             # серый — BASE без device_type
    return ORANGE


# ── Excel writer ──────────────────────────────────────────────────────────────

def write_audited_excel(source_path: Path, out_path: Path, vendor: str,
                         ai_predictions: dict[int, dict] | None = None):
    df_raw = pd.read_excel(source_path, header=None, dtype=str, engine="openpyxl").fillna("")

    # Find header row
    header_row_idx = None
    for i, row in df_raw.iterrows():
        vals = [str(v).strip().lower() for v in row.values]
        if "entity type" in vals or "entity_type" in vals:
            header_row_idx = i
            break
    if header_row_idx is None:
        print(f"  ⚠ Не найден заголовок в {source_path.name}, пропуск")
        return False, [], []

    df = df_raw.copy()
    raw_cols = [str(v).strip() for v in df_raw.iloc[header_row_idx].values]
    seen_cols: dict = {}
    deduped_cols = []
    for c in raw_cols:
        lc = c.lower().replace(" ", "_")
        if lc in seen_cols:
            seen_cols[lc] += 1
            deduped_cols.append(f"{c}_{seen_cols[lc]}")
        else:
            seen_cols[lc] = 0
            deduped_cols.append(c)
    df.columns = deduped_cols
    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
    col_map = {c: c.lower().replace(" ", "_") for c in df.columns}
    df_work = df.rename(columns=col_map)

    # Build data rows, normalizing vendor-specific column aliases
    _ALIASES = {"description": "option_name",
                "product_description": "option_name",
                "part_number": "skus",
                "product_#": "skus",
                "config_name": "module_name"}
    data_rows = []
    for row in df_work.to_dict("records"):
        for src, dst in _ALIASES.items():
            if src in row and dst not in row:
                row[dst] = row[src]
        data_rows.append(row)

    # Run rule-based validation + merge AI predictions
    all_results = []
    for i, row in enumerate(data_rows):
        rule_issues = validate_row(row, vendor, _device_type_map)

        ai_str = None
        if ai_predictions is not None:
            pred = ai_predictions.get(i)
            if pred:
                entity = str(row.get("entity_type") or "").strip().upper()
                device = str(row.get("device_type") or "").strip().lower()
                ai_str = build_ai_mismatch(entity, device, pred)

        parts = []
        if rule_issues:
            parts.extend(rule_issues)
        if ai_str:
            parts.append(ai_str)

        all_results.append("; ".join(parts) if parts else "OK")

    option_names = [str(row.get("option_name") or "").strip() for row in data_rows]

    # Write to Excel
    wb = openpyxl.load_workbook(source_path)
    ws = wb.active

    xlsx_header_row = header_row_idx + 1
    check_col_idx = ws.max_column + 1
    check_col_letter = get_column_letter(check_col_idx)

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    hdr = ws.cell(row=xlsx_header_row, column=check_col_idx, value="pipeline_check")
    hdr.fill = PatternFill("solid", fgColor="2C3E50")
    hdr.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    hdr.border = border

    data_start = xlsx_header_row + 1
    for i, result_str in enumerate(all_results):
        cell = ws.cell(row=data_start + i, column=check_col_idx, value=result_str)
        cell.font = Font(name="Arial", size=9)
        cell.border = border
        cell.alignment = Alignment(wrap_text=True)
        cell.fill = PatternFill("solid", fgColor=issue_color(result_str))

    ws.column_dimensions[check_col_letter].width = 60

    _write_summary_sheet(wb, all_results, vendor, source_path.name,
                          has_ai=ai_predictions is not None)
    wb.save(out_path)
    return True, all_results, option_names


def _write_summary_sheet(wb, all_results: list[str], vendor: str,
                          filename: str, has_ai: bool = False):
    if "Audit Summary" in wb.sheetnames:
        del wb["Audit Summary"]
    ws = wb.create_sheet("Audit Summary")

    thin = Side(border_style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_fill = PatternFill("solid", fgColor="2C3E50")
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)

    ai_label = "✅ с AI-проверкой" if has_ai else "⚡ только rule-checks"
    ws["A1"].value = f"Pipeline Audit — {filename}"
    ws["A1"].font = Font(bold=True, size=13, name="Arial", color="2C3E50")
    ws["A2"].value = f"Vendor: {vendor.upper()}  |  Mode: {ai_label}  |  {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = Font(italic=True, size=9, name="Arial", color="666666")

    total = len(all_results)
    ok_count = sum(1 for x in all_results if x == "OK")
    issue_rows = [x for x in all_results if x != "OK"]

    counts = [
        ("Всего строк (ITEM)", total, None),
        ("✅ OK", ok_count, GREEN),
        ("⚠ С проблемами", len(issue_rows), RED if issue_rows else GREEN),
        ("", "", None),
        ("🔴 E2 — UNKNOWN (нет правила)", sum(1 for x in issue_rows if "E2:" in x), RED),
        ("🟡 E4 — State mismatch", sum(1 for x in issue_rows if "E4:" in x), YELLOW),
        ("🟠 E7 — hw_type вне vocab", sum(1 for x in issue_rows if "E7:" in x), ORANGE),
        ("🟠 E8 — HW без hw_type", sum(1 for x in issue_rows if "E8:" in x), ORANGE),
        ("🟠 E9 — device_type→hw_type mismatch", sum(1 for x in issue_rows if "E9:" in x), ORANGE),
        ("🟠 E13 — LOGISTIC с физическим кабелем", sum(1 for x in issue_rows if "E13:" in x), ORANGE),
        ("🟠 E14 — CONFIG похож на blank_filler", sum(1 for x in issue_rows if "E14:" in x), ORANGE),
        ("⚪ E15 — BASE без device_type (info)", sum(1 for x in all_results if "E15:" in x), "F5F5F5"),
        ("🟣 E5/6/10/11/12 — Логика entity", sum(1 for x in issue_rows if any(f"E{n}:" in x for n in [5,6,10,11,12])), PURPLE),
    ]
    if has_ai:
        counts.append(("🔵 AI_MISMATCH — AI не согласен с пайплайном",
                        sum(1 for x in issue_rows if "AI_MISMATCH" in x), TEAL))
        counts.append(("🟠 MANUAL_CHECK — требует ручной проверки",
                        sum(1 for x in issue_rows if "MANUAL_CHECK" in x), "FFE0B2"))
        counts.append(("🟢 AI_SUGGEST — тип не определён, AI предлагает",
                        sum(1 for x in all_results if "AI_SUGGEST" in x), "E8F5E9"))
        counts.append(("⚪ E17 — HW строка без определённого типа",
                        sum(1 for x in all_results if "E17:" in x), "FAFAFA"))
        counts.append(("🟡 E18 — LOGISTIC с физическим keyword, нет device_type",
                        sum(1 for x in all_results if "E18:" in x), YELLOW))
        counts.append(("💛 E16 — заглушка в слоте (диска нет)",
                        sum(1 for x in all_results if "E16:" in x), "FFF9C4"))

    r = 4
    for label, val, color in counts:
        if label == "":
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = Font(name="Arial", size=10)
        ws.cell(row=r, column=1).border = border
        c = ws.cell(row=r, column=2, value=val)
        c.font = Font(name="Arial", size=10, bold=(isinstance(val, int) and val > 0))
        c.border = border
        if color and isinstance(val, int) and val > 0:
            c.fill = PatternFill("solid", fgColor=color)
        r += 1

    # Legend
    r += 2
    ws.cell(row=r, column=1, value="Легенда").font = Font(bold=True, size=11, name="Arial")
    r += 1
    legend = [
        (RED,    "🔴 UNKNOWN — строка не распознана ни одним правилом"),
        (TEAL,   "🔵 AI_MISMATCH — AI не согласен с пайплайном"),
        ("FFE0B2","🟠 MANUAL_CHECK — AI неуверен, нужна ручная проверка"),
        ("E8F5E9","🟢 AI_SUGGEST — пайплайн не определил тип, AI предлагает вариант"),
        (PURPLE, "🟣 Логика entity (hw_type на BASE/CONFIG/NOTE)"),
        (ORANGE, "🟠 hw_type проблемы (E7/E8/E9/E13/E14)"),
        (YELLOW, "🟡 State mismatch (E4)"),
        ("F5F5F5","⚪ Info — BASE без device_type (E15, не ошибка)"),
        (GREEN,  "✅ OK — все проверки пройдены"),
    ]
    for color, desc in legend:
        c1 = ws.cell(row=r, column=1, value="  ")
        c1.fill = PatternFill("solid", fgColor=color)
        c1.border = border
        c2 = ws.cell(row=r, column=2, value=desc)
        c2.font = Font(name="Arial", size=9)
        c2.border = border
        r += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 22


# ── File discovery ────────────────────────────────────────────────────────────

def _tag_to_comment(tag: str) -> str:
    """Convert pipeline_check tag to human-readable Russian comment."""
    import re
    t = tag.strip()

    # AI_MISMATCH entity
    m = re.search(r'AI_MISMATCH.*:entity\[pipeline:(\w+)→ai:(\w+)\]', t)
    if m:
        labels = {"HW": "железо", "CONFIG": "конфигурация", "SOFTWARE": "ПО",
                  "LOGISTIC": "логистика", "NOTE": "примечание", "BASE": "базовый продукт",
                  "SERVICE": "сервис", "UNKNOWN": "неизвестно"}
        p = labels.get(m.group(1), m.group(1))
        a = labels.get(m.group(2), m.group(2))
        return f"AI считает тип '{a}', пайплайн поставил '{p}' — проверить"

    # AI_MISMATCH device_type
    m = re.search(r'AI_MISMATCH.*:device_type\[pipeline:(\w+)→ai:(\w+)\]', t)
    if m:
        return f"AI считает device_type='{m.group(2)}', пайплайн поставил '{m.group(1)}' — возможная ошибка классификации"

    # AI_SUGGEST
    m = re.search(r'AI_SUGGEST:device_type=(\w+)', t)
    if m:
        return f"Пайплайн не определил тип устройства. AI предлагает: {m.group(1)}"

    # MANUAL_CHECK
    m = re.search(r'MANUAL_CHECK:device_type_unknown\[ai_guess:(\w+)\]', t)
    if m:
        return f"Тип устройства не определён. AI неуверен, возможно: {m.group(1)} — нужна ручная проверка"

    # E-codes
    if "E13:" in t:
        m = re.search(r'device:(\w+)', t)
        dev = m.group(1) if m else "кабель/шнур"
        return f"Тип LOGISTIC для физического компонента '{dev}' — возможно должен быть HW"
    if "E14:" in t:
        return "CONFIG-строка похожа на заглушку (blank_filler) — возможно пропущен device_type"
    if "E15:" in t:
        return "BASE-строка без device_type — базовый продукт, это нормально"
    if "E16:" in t:
        return "Заглушка в слоте (blank_filler), диск отсутствует — слот физически занят заглушкой"
    if "E17:" in t:
        return "HW-строка без определённого типа устройства — пайплайн не смог классифицировать"
    if "E2:" in t:
        return "Нет подходящего правила — entity_type=UNKNOWN"
    if "E9:" in t:
        m = re.search(r'device:(\w+),expected:(\w+),got:(\w+)', t)
        if m:
            return f"Несоответствие: device_type='{m.group(1)}' ожидает hw_type='{m.group(2)}', но стоит '{m.group(3)}'"
    return t  # fallback — original tag


def _generate_human_report(report_files: list, output_dir: str, llm_model: str) -> None:
    """Generate audit_summary.xlsx — human-readable report with all problem rows."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводный отчёт"

    # ── Header ────────────────────────────────────────────────────────────────
    headers = [
        "Документ", "Вендор", "Артикул / SKU", "Описание", "Module Name",
        "Entity Type", "HW Type", "Device Type", "State",
        "Пометка скрипта", "Комментарий"
    ]
    col_widths = [22, 8, 18, 55, 22, 12, 16, 16, 10, 25, 55]

    header_fill = PatternFill("solid", fgColor="1F3864")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    # ── Color map for tags ────────────────────────────────────────────────────
    tag_colors = {
        "AI_MISMATCH": "DDEEFF",
        "MANUAL_CHECK": "FFE0B2",
        "AI_SUGGEST":   "E8F5E9",
        "E13":          "FFE0CC",
        "E14":          "FFE0CC",
        "E16":          "FFF9C4",
        "E17":          "FAFAFA",
        "E15":          "F5F5F5",
        "E2":           "FFCCCC",
        "E9":           "FFE0CC",
    }

    def row_color(tag: str) -> str:
        for key, color in tag_colors.items():
            if key in tag:
                return color
        return "FFFFFF"

    # ── Read each audited file and collect problem rows ───────────────────────
    row_num = 2
    total_written = 0

    for fdata in report_files:
        # Find the audited file path
        # results contains pipeline_check values, but we need full row data
        # Re-read the audited xlsx
        import openpyxl as _opx
        from pathlib import Path

        # Find audited file — same dir as original, with _audited suffix
        orig_name = fdata["file"]
        stem = Path(orig_name).stem
        # Search in output_dir recursively
        audited_files = list(Path(output_dir).rglob(f"{stem}_audited.xlsx"))
        if not audited_files:
            continue
        audited_path = audited_files[0]

        try:
            awb = _opx.load_workbook(audited_path, read_only=True, data_only=True)
            aws = awb.active

            # Find header row and columns
            hdr_row_idx = None
            cols = {}
            for i, row in enumerate(aws.iter_rows(values_only=True)):
                if row and any(str(v).strip().lower() in ("entity type","entity_type") for v in row if v):
                    hdr_row_idx = i
                    for j, v in enumerate(row):
                        sv = str(v).strip().lower().replace(" ", "_") if v else ""
                        if sv == "pipeline_check": cols["check"] = j
                        if sv in ("entity_type","entity type"): cols["entity"] = j
                        if sv == "device_type": cols["device"] = j
                        if sv == "hw_type": cols["hw"] = j
                        if sv in ("option_name","description","product_description"): cols["option"] = j
                        if sv in ("module_name","module name","config_name"): cols["module"] = j
                        if sv in ("skus","sku","part_number","product_#"): cols["sku"] = j
                        if sv == "state": cols["state"] = j
                    break

            if "check" not in cols:
                awb.close()
                continue

            def gcell(row, key):
                idx = cols.get(key)
                if idx is None or idx >= len(row):
                    return ""
                return str(row[idx] or "").strip()

            for row in aws.iter_rows(min_row=(hdr_row_idx or 0)+2, values_only=True):
                if not row:
                    continue
                check_val = gcell(row, "check")
                if not check_val or check_val == "OK":
                    continue

                # Build comment from all tags
                tags = [t.strip() for t in check_val.split(";") if t.strip()]
                comments = [_tag_to_comment(t) for t in tags]
                comment_str = " | ".join(comments)
                tag_str = "; ".join(tags)

                color = row_color(check_val)
                fill = PatternFill("solid", fgColor=color)
                data_font = Font(name="Arial", size=9)
                wrap = Alignment(wrap_text=True, vertical="top")

                values = [
                    orig_name,
                    fdata["vendor"].upper(),
                    gcell(row, "sku"),
                    gcell(row, "option"),
                    gcell(row, "module"),
                    gcell(row, "entity"),
                    gcell(row, "hw"),
                    gcell(row, "device"),
                    gcell(row, "state"),
                    tag_str,
                    comment_str,
                ]

                for col, val in enumerate(values, 1):
                    cell = ws.cell(row=row_num, column=col, value=val)
                    cell.font = data_font
                    cell.fill = fill
                    cell.alignment = wrap
                    cell.border = border

                ws.row_dimensions[row_num].height = 30
                row_num += 1
                total_written += 1

            awb.close()
        except Exception as e:
            print(f"    ⚠ Human report: ошибка чтения {orig_name}: {e}")

    # ── Legend sheet ──────────────────────────────────────────────────────────
    ls = wb.create_sheet("Легенда")
    legend_data = [
        ("DDEEFF", "🔵 AI_MISMATCH — AI не согласен с классификацией пайплайна"),
        ("FFE0B2", "🟠 MANUAL_CHECK — AI неуверен, нужна ручная проверка"),
        ("E8F5E9", "🟢 AI_SUGGEST — пайплайн не определил тип, AI предлагает вариант"),
        ("FFE0CC", "🟠 E13/E14 — логика entity/device требует внимания"),
        ("FFCCCC", "🔴 E2 — нет правила классификации (UNKNOWN)"),
        ("FFF9C4", "💛 E16 — заглушка в слоте, диска нет"),
        ("FAFAFA", "⚪ E17 — HW без определённого типа"),
        ("F5F5F5", "⚫ E15 — BASE без device_type (норма)"),
    ]
    ls.column_dimensions["A"].width = 5
    ls.column_dimensions["B"].width = 60
    for i, (color, desc) in enumerate(legend_data, 2):
        ls.cell(row=i, column=1).fill = PatternFill("solid", fgColor=color)
        ls.cell(row=i, column=2, value=desc).font = Font(name="Arial", size=10)

    # ── Save ──────────────────────────────────────────────────────────────────
    out_path = Path(output_dir) / "audit_summary.xlsx"
    wb.save(out_path)
    print(f"📊 Сводный отчёт: {out_path}  ({total_written} строк)")


# Known false-positive cases. Each entry must specify vendor + condition.
# Do NOT add broad transition-only patterns — they mask future vendor bugs.
KNOWN_FP_CASES = [
    # HPE Cable Kit: AI says accessory, but Cable Kit = cable per business rules.
    # Matches: "OCPA Cable Kit", "GPU Power Cable Kit", "NVMe to Tri-Mode OCP FIO Cable Kit"
    {
        "vendor": "hpe",
        "kind": "device_mismatch",
        "transition": "cable→accessory",
        "option_name_pattern": r"(?i)cable\s+(kit|assembly)",
    },
    # HPE Hybrid Capacitor: AI says accessory, but Capacitor = battery per business rules.
    # Matches: "HPE Smart Storage Hybrid Capacitor with 260mm/145mm Cable Kit"
    {
        "vendor": "hpe",
        "kind": "device_mismatch",
        "transition": "battery→accessory",
        "option_name_pattern": r"(?i)hybrid\s+capacitor",
    },
    # HPE Rail Kit: AI says accessory, but Rail Kit = rail per business rules.
    # Matches: "HPE ProLiant DL3XX Gen11 Easy Install Rail 3 Kit"
    {
        "vendor": "hpe",
        "kind": "device_mismatch",
        "transition": "rail→accessory",
        "option_name_pattern": r"(?i)(rail\s+kit|easy\s+install\s+rail)",
    },
    # HPE NVLink Bridge: GPU interconnect accessory, not a GPU itself.
    # Matches: "NVIDIA 4-way NVLink Bridge for H200 NVL"
    {
        "vendor": "hpe",
        "kind": "device_mismatch",
        "transition": "accessory→gpu",
        "option_name_pattern": r"(?i)nvlink\s+bridge",
    },
    # HPE Cable Management Arm: AI says cable, but it's a mounting accessory.
    # Matches: "HPE DL38X Gen10 Plus 2U Cable Management Arm for Rail Kit"
    {
        "vendor": "hpe",
        "kind": "device_mismatch",
        "transition": "accessory→cable",
        "option_name_pattern": r"(?i)cable\s+management\s+arm",
    },
    # HPE BASE server CTO rows: AI thinks NOTE/HW/LOGISTIC, but BASE is correct.
    # Matches: "HPE ProLiant DL380 Gen11 ... Configure-to-order Server"
    {
        "vendor": "hpe",
        "kind": "entity_mismatch",
        "transition": "BASE→NOTE",
        "option_name_pattern": r"(?i)(?:configure[- ]to[- ]order|cto)\s+(?:server|svr)\b",
    },
    {
        "vendor": "hpe",
        "kind": "entity_mismatch",
        "transition": "BASE→LOGISTIC",
        "option_name_pattern": r"(?i)(?:configure[- ]to[- ]order|cto)\s+(?:server|svr)\b",
    },
    # Dell FRONT STORAGE module: AI reads "storage_controller", but this is a chassis
    # option describing physical drive bay count (e.g. 379-BDTF). The PERC controller
    # always appears as a separate row.
    # Matches: option_name for rows where module_name="FRONT STORAGE"
    {
        "vendor": "dell",
        "kind": "device_mismatch",
        "transition": "chassis→storage_controller",
        "option_name_pattern": r"(?i)\bfront\s+storage\b",
    },
    # Dell Chassis Configuration + NVMe Hardware RAID Drives: AI reads
    # "storage_controller", but this is a chassis form-factor choice
    # (e.g. 321-BLMP, 750-BBPB). PERC 12 is always a separate row.
    {
        "vendor": "dell",
        "kind": "device_mismatch",
        "transition": "chassis→storage_controller",
        "option_name_pattern": r"(?i)\bNVMe\s+Hardware\s+RAID\b",
    },
]


def _is_known_fp(bug_items: list[dict], kind: str, transition: str) -> bool:
    """Check if ALL items in a bug pattern match a known FP case."""
    for case in KNOWN_FP_CASES:
        if case["kind"] != kind or case["transition"] != transition:
            continue
        pattern = re.compile(case["option_name_pattern"])
        vendor = case["vendor"]
        if all(
            i.get("vendor") == vendor
            and pattern.search(i.get("option_name", ""))
            for i in bug_items
        ):
            return True
    return False


def _generate_report(report_files: list, output_dir: str, llm_model: str,
                     tok_in: int, tok_out: int, use_ai: bool) -> None:
    """Generate audit_report.json next to OUTPUT dir."""
    import re
    from collections import Counter, defaultdict
    from datetime import datetime

    # ── Aggregate all results ─────────────────────────────────────────────────
    all_issues = []   # {file, vendor, tag, entity, device, hw, option}
    stats_by_vendor = defaultdict(lambda: {"files": 0, "rows": 0, "ok": 0, "issues": 0})
    tag_counter = Counter()

    for fdata in report_files:
        vendor = fdata["vendor"]
        stats_by_vendor[vendor]["files"] += 1
        stats_by_vendor[vendor]["rows"] += fdata["total_rows"]
        stats_by_vendor[vendor]["ok"] += fdata["ok"]
        stats_by_vendor[vendor]["issues"] += fdata["issues"]

    # Re-read audited files to get row-level detail
    for fdata in report_files:
        opt_names = fdata.get("option_names", [])
        for idx, val in enumerate(fdata.get("results", [])):
            if not val or val == "OK":
                continue
            opt_name = opt_names[idx] if idx < len(opt_names) else ""
            for tag in str(val).split(";"):
                tag = tag.strip()
                if not tag:
                    continue
                prefix = tag.split(":")[0].strip()
                tag_counter[prefix] += 1
                all_issues.append({
                    "file": fdata["file"], "vendor": fdata["vendor"],
                    "tag": tag, "option_name": opt_name,
                })

    # ── Real bugs: AI_MISMATCH patterns ──────────────────────────────────────
    mismatch_groups = defaultdict(list)
    for issue in all_issues:
        tag = issue["tag"]
        if "AI_MISMATCH" not in tag:
            continue
        # entity mismatch
        m = re.search(r'entity\[pipeline:(\w+)→ai:(\w+)\]', tag)
        if m:
            key = f"entity_mismatch:{m.group(1)}→{m.group(2)}"
            mismatch_groups[key].append(issue)
        # device_type mismatch
        m = re.search(r'device_type\[pipeline:(\w+)→ai:(\w+)\]', tag)
        if m:
            key = f"device_mismatch:{m.group(1)}→{m.group(2)}"
            mismatch_groups[key].append(issue)

    bugs = []
    for pattern, items in sorted(mismatch_groups.items(), key=lambda x: -len(x[1])):
        parts = pattern.split(":")
        kind = parts[0]
        transition = parts[1] if len(parts) > 1 else ""
        from_val, to_val = transition.split("→") if "→" in transition else ("", "")
        
        # NOTE: HPE device_mismatch patterns (server→cable, server→fan, server→heatsink, etc.)
        # are NOT in fp_patterns — they appear as REVIEW_NEEDED or REAL_BUG intentionally.
        # HPE pipeline sets device_type=server on component rows (known HPE rule behaviour).
        # Do not add server→* to fp_patterns without confirming HPE rules change.
        fp_patterns = {
            "HW→LOGISTIC", "HW→NOTE", "HW→CONFIG", "SOFTWARE→NOTE",
            "LOGISTIC→NOTE", "SERVICE→NOTE", "CONFIG→NOTE", "BASE→CONFIG"
        }

        unique_files = len(set(i["file"] for i in items))
        transition = f"{from_val}→{to_val}"
        if _is_known_fp(items, kind, transition):
            bug_type = "FALSE_POSITIVE"
        elif kind == "entity_mismatch" and transition in fp_patterns:
            bug_type = "FALSE_POSITIVE"
        elif kind == "device_mismatch":
            if len(items) >= 3 or (len(items) >= 2 and unique_files >= 2):
                bug_type = "REAL_BUG"
            else:
                bug_type = "REVIEW_NEEDED"
        elif kind == "entity_mismatch":
            bug_type = "REVIEW_NEEDED"
        else:
            bug_type = "REVIEW_NEEDED"

        bugs.append({
            "type": bug_type,
            "pattern": pattern,
            "count": len(items),
            "vendors": list(set(i["vendor"] for i in items)),
            "examples": [{"file": i["file"], "tag": i["tag"]} for i in items[:3]],
            "fix_target": "rules/{vendor}_rules.yaml" if bug_type == "REAL_BUG" else None,
        })

    # ── AI_SUGGEST: yaml candidates ──────────────────────────────────────────
    suggest_groups = defaultdict(list)
    for issue in all_issues:
        if "AI_SUGGEST" not in issue["tag"]:
            continue
        m = re.search(r'device_type=(\w+)', issue["tag"])
        if m:
            suggest_groups[m.group(1)].append(issue)

    yaml_candidates = []
    for device_type, items in sorted(suggest_groups.items(), key=lambda x: -len(x[1])):
        yaml_candidates.append({
            "device_type": device_type,
            "count": len(items),
            "vendors": list(set(i["vendor"] for i in items)),
            "note": "pipeline не проставил device_type, AI уверенно предлагает",
        })

    # ── E-code breakdown ─────────────────────────────────────────────────────
    e_groups = defaultdict(list)
    for issue in all_issues:
        tag = issue["tag"]
        if tag.startswith("E"):
            code = tag.split(":")[0]
            e_groups[code].append(issue)

    rule_issues = [
        {"code": code, "count": len(items),
         "vendors": list(set(i["vendor"] for i in items)),
         "examples": [i["tag"] for i in items[:3]]}
        for code, items in sorted(e_groups.items(), key=lambda x: -len(x[1]))
    ]

    # ── Pricing ───────────────────────────────────────────────────────────────
    pricing = PRICING.get(llm_model, {"in": 0, "out": 0})
    total_cost = (tok_in * pricing["in"] + tok_out * pricing["out"]) / 1_000_000

    # ── Assemble report ───────────────────────────────────────────────────────
    report = {
        "meta": {
            "run_date": datetime.now().strftime("%Y-%m-%d %H:%M"),
            "model": llm_model if use_ai else "no-ai",
            "total_tokens": tok_in + tok_out,
            "tokens_in": tok_in,
            "tokens_out": tok_out,
            "cost_usd": round(total_cost, 4),
        },
        "stats": {
            "total_files": len(report_files),
            "total_rows": sum(f["total_rows"] for f in report_files),
            "ok": sum(f["ok"] for f in report_files),
            "issues": sum(f["issues"] for f in report_files),
            "by_tag": dict(tag_counter.most_common()),
            "by_vendor": {v: dict(s) for v, s in stats_by_vendor.items()},
            "by_file": [
                {"file": f["file"], "vendor": f["vendor"],
                 "rows": f["total_rows"], "ok": f["ok"],
                 "issues": f["issues"], "cost_usd": f["cost_usd"]}
                for f in report_files
            ],
        },
        "bugs": sorted(bugs, key=lambda x: (x["type"] != "REAL_BUG", -x["count"])),
        "yaml_candidates": yaml_candidates,
        "rule_issues": rule_issues,
        "claude_prompt": _build_claude_prompt(bugs, yaml_candidates, rule_issues),
    }

    # ── Save ──────────────────────────────────────────────────────────────────
    out_path = Path(output_dir) / "audit_report.json"
    with open(out_path, "w", encoding="utf-8") as fh:
        json.dump(report, fh, ensure_ascii=False, indent=2)
    print(f"\n📄 Отчёт сохранён: {out_path}")


def _build_claude_prompt(bugs: list, yaml_candidates: list, rule_issues: list) -> str:
    """Build a ready-to-use prompt for Claude/Cursor to fix Teresa rules."""
    lines = [
        "Ты помогаешь улучшить систему классификации оборудования Teresa.",
        "Ниже — результаты аудита пайплайна. Подготовь конкретные правки для YAML-файлов правил.",
        "",
        "## РЕАЛЬНЫЕ БАГИ (требуют правок в YAML):",
    ]
    real_bugs = [b for b in bugs if b["type"] == "REAL_BUG"]
    if real_bugs:
        for bug in real_bugs:
            lines.append(f"- {bug['pattern']} × {bug['count']} строк у вендоров {bug['vendors']}")
            for ex in bug.get("examples", [])[:2]:
                lines.append(f"  Пример: {ex['tag']}")
    else:
        lines.append("- Нет явных багов")

    lines += [
        "",
        "## КАНДИДАТЫ НА НОВЫЕ ПРАВИЛА (AI_SUGGEST — пайплайн не определил device_type):",
    ]
    for cand in yaml_candidates[:10]:
        lines.append(f"- device_type={cand['device_type']} × {cand['count']} строк у {cand['vendors']}")

    lines += [
        "",
        "## ТРЕБУЕТ РЕВЬЮ (спорные классификации):",
    ]
    review = [b for b in bugs if b["type"] == "REVIEW_NEEDED"]
    for bug in review[:5]:
        lines.append(f"- {bug['pattern']} × {bug['count']} строк")

    lines += [
        "",
        "## ЗАДАЧА:",
        "1. Для каждого РЕАЛЬНОГО БАГА — предложи конкретное regex-правило для YAML",
        "2. Для AI_SUGGEST кандидатов — предложи новые правила device_type в YAML",
        "3. Для спорных случаев — объясни почему пайплайн прав или неправ",
        "4. Выдай итоговый список правок в формате: файл → что добавить/изменить",
    ]
    return "\n".join(lines)


def find_annotated_files(output_dir: Path, vendor_filter: str | None,
                          since: str | None) -> list[Path]:
    files = sorted(output_dir.rglob("*_annotated.xlsx"))
    # Exclude already-audited files
    files = [f for f in files if "_audited" not in f.name]
    # Exclude TOTAL aggregation folders (contain duplicates of per-run files)
    files = [f for f in files if "-TOTAL" not in f.parent.name]
    if vendor_filter:
        files = [f for f in files if vendor_filter in str(f).lower()]
    if since:
        cutoff = datetime.strptime(since, "%Y-%m-%d")
        files = [f for f in files if datetime.fromtimestamp(f.stat().st_mtime) >= cutoff]
    return files


def detect_vendor_from_path(path: Path, known_vendors: list[str] | None = None) -> str:
    """Detect vendor from path components using known vendor list."""
    if known_vendors is None:
        known_vendors = _get_known_vendors(_load_config())
    s = str(path).lower()
    for vendor in known_vendors:
        if f"{vendor}_run" in s or f"/{vendor}/" in s or f"\\{vendor}\\" in s:
            return vendor
    # HPE alias: "hp_run" → "hpe"
    if "hp_run" in s:
        return "hpe"
    print(f"  [WARN] Cannot detect vendor from path: {path}", file=sys.stderr)
    return "unknown"


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    config = _load_config()
    known_vendors = _get_known_vendors(config)

    parser = argparse.ArgumentParser(description="Batch pipeline audit v2 (rule + AI)")
    parser.add_argument("--output-dir", default=".", help="Root OUTPUT folder")
    parser.add_argument("--vendor", choices=known_vendors)
    parser.add_argument("--since", metavar="YYYY-MM-DD")
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--no-ai", action="store_true", help="Skip LLM predictions (faster)")
    parser.add_argument("--suffix", default="_audited")
    parser.add_argument("--batch-size", type=int, default=40,
                        help="Rows per LLM batch (default: 40)")
    parser.add_argument("--provider", choices=["anthropic", "openai"], default="openai",
                        help="LLM provider (default: openai)")
    parser.add_argument("--model", default="gpt-4o-mini",
                        help="Model name (default: gpt-4o-mini)")
    args = parser.parse_args()

    output_dir = Path(args.output_dir)
    if not output_dir.exists():
        print(f"❌ Папка не найдена: {output_dir}")
        sys.exit(1)

    files = find_annotated_files(output_dir, args.vendor, args.since)
    if not files:
        print("Файлы *_annotated.xlsx не найдены.")
        sys.exit(0)

    print(f"Найдено {len(files)} файлов:")
    for f in files:
        print(f"  {f.relative_to(output_dir)}")

    if args.dry_run:
        print("\n[dry-run] Обработка не запускалась.")
        return

    # Init Anthropic client if AI enabled
    client = None
    use_ai = not args.no_ai
    llm_model = args.model
    if use_ai:
        provider = args.provider
        try:
            if provider == "openai":
                import openai
                api_key = os.environ.get("OPENAI_API_KEY")
                if not api_key:
                    print("⚠ OPENAI_API_KEY не найден → запуск без AI-проверки")
                    print("  Чтобы включить: $env:OPENAI_API_KEY=\"sk-...\"")
                    use_ai = False
                else:
                    client = openai.OpenAI(api_key=api_key)
                    print(f"🤖 AI-режим: OpenAI ({llm_model}, batch={args.batch_size})")
            else:
                import anthropic
                api_key = os.environ.get("ANTHROPIC_API_KEY")
                if not api_key:
                    print("⚠ ANTHROPIC_API_KEY не найден → запуск без AI-проверки")
                    print("  Чтобы включить: $env:ANTHROPIC_API_KEY=\"sk-ant-...\"")
                    use_ai = False
                else:
                    client = anthropic.Anthropic(api_key=api_key)
                    print(f"🤖 AI-режим: Anthropic ({llm_model}, batch={args.batch_size})")
        except ImportError as e:
            print(f"⚠ Библиотека не установлена ({e}) → pip install openai  или  pip install anthropic")
            print("  Запуск без AI-проверки.")
            use_ai = False

    print()
    ok_count, failed = 0, 0
    session_tok_in = 0
    session_tok_out = 0
    report_files = []

    for f in files:
        vendor = args.vendor or detect_vendor_from_path(f, known_vendors)
        out_path = f.parent / f"{f.stem}{args.suffix}.xlsx"
        file_num = files.index(f) + 1
        print(f"  [{file_num}/{len(files)}] [{vendor.upper()}] {f.name}", flush=True)

        ai_predictions = None
        f_tok_in = 0
        f_tok_out = 0
        if use_ai and client:
            # Load rows for LLM
            try:
                df_raw = pd.read_excel(f, header=None, dtype=str, engine="openpyxl").fillna("")
                header_row_idx = None
                for i, row in df_raw.iterrows():
                    vals = [str(v).strip().lower() for v in row.values]
                    if "entity type" in vals or "entity_type" in vals:
                        header_row_idx = i
                        break
                if header_row_idx is not None:
                    df = df_raw.copy()
                    raw_cols2 = [str(v).strip() for v in df_raw.iloc[header_row_idx].values]
                    seen2: dict = {}
                    deduped2 = []
                    for c in raw_cols2:
                        lc2 = c.lower().replace(" ", "_")
                        if lc2 in seen2:
                            seen2[lc2] += 1
                            deduped2.append(f"{c}_{seen2[lc2]}")
                        else:
                            seen2[lc2] = 0
                            deduped2.append(c)
                    df.columns = deduped2
                    df = df.iloc[header_row_idx + 1:].reset_index(drop=True)
                    col_map = {c: c.lower().replace(" ", "_") for c in df.columns}
                    df_work = df.rename(columns=col_map)
                    _AL = {"description": "option_name", "product_description": "option_name", "part_number": "skus", "product_#": "skus"}
                    data_rows = []
                    for row in df_work.to_dict("records"):
                        for src, dst in _AL.items():
                            if src in row and dst not in row:
                                row[dst] = row[src]
                        data_rows.append(row)
                    ai_predictions, f_tok_in, f_tok_out = run_llm_predictions(data_rows, client, args.batch_size, llm_model)
                    session_tok_in += f_tok_in
                    session_tok_out += f_tok_out
            except Exception as e:
                print(f"    ⚠ LLM prep error: {e}")

        try:
            success, results, option_names = write_audited_excel(f, out_path, vendor, ai_predictions)
            if success:
                issue_count = sum(1 for r in results if r != "OK")
                ai_count = sum(1 for r in results if "AI_MISMATCH" in r)
                ai_tag = f", {ai_count} AI mismatch" if use_ai else ""
                # Per-file token cost
                cost_str = ""
                file_cost = 0.0
                if use_ai and (f_tok_in + f_tok_out) > 0:
                    pricing = PRICING.get(llm_model, {"in": 0, "out": 0})
                    file_cost = (f_tok_in * pricing["in"] + f_tok_out * pricing["out"]) / 1_000_000
                    cost_str = f"  |  {f_tok_in+f_tok_out:,} токенов  ${file_cost:.4f}"
                print(f"    → {out_path.name} ✅  проблем: {issue_count}{ai_tag}{cost_str}")
                ok_count += 1
                # Collect for report
                report_files.append({
                    "file": f.name,
                    "vendor": vendor,
                    "total_rows": len(results),
                    "ok": sum(1 for r in results if r == "OK"),
                    "issues": issue_count,
                    "tokens": f_tok_in + f_tok_out,
                    "cost_usd": round(file_cost, 4),
                    "results": results,
                    "option_names": option_names,
                })
            else:
                failed += 1
        except Exception as e:
            print(f"    ❌ {e}")
            failed += 1

    print(f"\nГотово: {ok_count} обработано, {failed} ошибок.")
    if use_ai and (session_tok_in + session_tok_out) > 0:
        pricing = PRICING.get(llm_model, {"in": 0, "out": 0})
        total_cost = (session_tok_in * pricing["in"] + session_tok_out * pricing["out"]) / 1_000_000
        print(f"💰 Итого токенов: {session_tok_in+session_tok_out:,}  (in: {session_tok_in:,} / out: {session_tok_out:,})")
        print(f"💰 Стоимость прогона: ${total_cost:.4f}  (модель: {llm_model})")
    if use_ai:
        print("Файлы содержат колонку pipeline_check с AI_MISMATCH пометками.")
    print(f"Сохранены рядом с исходными как *{args.suffix}.xlsx")

    # ── Generate JSON report ──────────────────────────────────────────────────
    if report_files:
        _generate_report(report_files, args.output_dir, llm_model,
                         session_tok_in, session_tok_out, use_ai)
        _generate_human_report(report_files, args.output_dir, llm_model)


if __name__ == "__main__":
    main()
