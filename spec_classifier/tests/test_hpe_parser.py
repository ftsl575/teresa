"""Unit tests for HPE BOM parser (in-memory workbooks)."""

import pytest
import openpyxl

from src.vendors.hpe.parser import parse_excel


def test_parse_hpe_valid_bom_and_row_index(tmp_path):
    """Valid BOM sheet: parses rows and preserves 1-based Excel row index."""
    path = tmp_path / "hpe_valid.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    ws.append(["Product #", "Product Description", "Qty", "Unit Price (USD)", "Config Name"])
    ws.append(["P123 ABC", "Intel Xeon CPU", 2, 1500.0, "Server A"])
    ws.append([None, None, None, None, None])  # fully empty data row should be skipped
    ws.append(["X999", "16GB DDR5", 4, 120.0, "Server A"])
    ws.append(["Total", None, None, None, None])  # end-of-data marker
    wb.save(path)
    wb.close()

    rows, header_idx = parse_excel(str(path))
    assert header_idx == 0
    assert len(rows) == 2
    assert rows[0]["Product #"] == "P123 ABC"
    assert rows[0]["__row_index__"] == 2
    assert rows[1]["Product Description"] == "16GB DDR5"
    assert rows[1]["__row_index__"] == 4


def test_parse_hpe_empty_bom_returns_empty(tmp_path):
    """BOM exists but has no meaningful rows -> empty result."""
    path = tmp_path / "hpe_empty.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BOM"
    wb.save(path)
    wb.close()

    rows, header_idx = parse_excel(str(path))
    assert rows == []
    assert header_idx == 0


def test_parse_hpe_missing_bom_sheet_raises(tmp_path):
    """Workbook without BOM sheet should raise ValueError."""
    path = tmp_path / "hpe_no_bom.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A", "B"])
    wb.save(path)
    wb.close()

    with pytest.raises(ValueError, match="Sheet 'BOM' not found"):
        parse_excel(str(path))


def test_parse_hpe_file_not_found_raises():
    """Non-existent path should raise FileNotFoundError."""
    with pytest.raises(FileNotFoundError):
        parse_excel("Z:/definitely/missing/file.xlsx")
