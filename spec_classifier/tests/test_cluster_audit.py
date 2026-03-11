"""Tests for cluster_audit.py — public function coverage."""

import json
import math
import sys
from pathlib import Path

import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
import openpyxl
import pandas as pd

from cluster_audit import (
    _detect_vendor_from_path,
    _is_empty,
    _collect_xlsx_files,
    _load_xlsx,
    load_candidate_rows,
    normalize_text,
    analyze_clusters,
    heuristic_mapping,
    write_cluster_summary,
    print_dry_run_report,
    build_parser,
)


def _make_xlsx(path, columns, rows):
    """Create a minimal xlsx file for testing."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(columns)
    for row in rows:
        ws.append(row)
    wb.save(path)
    wb.close()


# ---------------------------------------------------------------------------
# _detect_vendor_from_path
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("path_str, expected", [
    # vendor in stem
    ("OUTPUT/hpe_run/hp8_annotated_audited.xlsx", "hpe"),
    ("OUTPUT/dell_run/dl5_annotated_audited.xlsx", "dell"),
    ("OUTPUT/cisco_run/ccw_2_annotated_audited.xlsx", "cisco"),
    # vendor in parent directory name
    ("OUTPUT/hpe/some_file.xlsx", "hpe"),
    ("OUTPUT/dell/some_file.xlsx", "dell"),
    # nothing matches → unknown
    ("OUTPUT/other/mystery_file.xlsx", "unknown"),
])
def test_detect_vendor_from_path(path_str, expected):
    assert _detect_vendor_from_path(Path(path_str)) == expected


# ---------------------------------------------------------------------------
# _is_empty
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("val, expected", [
    (None,        True),
    (math.nan,    True),
    ("",          True),
    ("   ",       True),
    ("hello",     False),
    (0,           False),    # 0 is not empty — str(0) == "0"
    ("0",         False),
])
def test_is_empty(val, expected):
    assert _is_empty(val) is expected


# ---------------------------------------------------------------------------
# _collect_xlsx_files — prefer *_annotated_audited over *_annotated
# ---------------------------------------------------------------------------

def test_collect_xlsx_files_prefers_audited(tmp_path):
    """When both *_annotated.xlsx and *_annotated_audited.xlsx exist for the
    same source stem, only the audited variant should be returned."""
    (tmp_path / "widget_annotated.xlsx").touch()
    (tmp_path / "widget_annotated_audited.xlsx").touch()

    result = _collect_xlsx_files(tmp_path)
    names = [p.name for p in result]

    assert "widget_annotated_audited.xlsx" in names
    assert "widget_annotated.xlsx" not in names


def test_collect_xlsx_files_includes_annotated_only(tmp_path):
    """When only *_annotated.xlsx exists (no audited), it should be returned."""
    (tmp_path / "gadget_annotated.xlsx").touch()

    result = _collect_xlsx_files(tmp_path)
    names = [p.name for p in result]
    assert "gadget_annotated.xlsx" in names


def test_collect_xlsx_files_empty_dir(tmp_path):
    assert _collect_xlsx_files(tmp_path) == []


# ---------------------------------------------------------------------------
# normalize_text
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("text, checks", [
    # lowercasing
    ("Intel XEON Gold", lambda t: t == t.lower()),
    # dollar removal
    ("Kit $1,234.00 upgrade", lambda t: "$" not in t and "1,234" not in t),
    # SFP+ normalisation
    ("SFP+ transceiver 10G", lambda t: "sfp_plus" in t),
    # extra whitespace collapsed
    ("some   option   name", lambda t: "  " not in t),
])
def test_normalize_text(text, checks):
    result = normalize_text(text)
    assert checks(result)


# ---------------------------------------------------------------------------
# analyze_clusters
# ---------------------------------------------------------------------------

def test_analyze_clusters_skips_noise_label():
    """Rows assigned label -1 (HDBSCAN noise) must not appear in output."""
    rows = [
        {"option_name": "NVMe SSD",  "vendor": "hpe", "source_file": "f.xlsx"},
        {"option_name": "CPU Xeon",  "vendor": "hpe", "source_file": "f.xlsx"},
        {"option_name": "Fan Kit",   "vendor": "hpe", "source_file": "f.xlsx"},
    ]
    labels = [-1, 0, 0]
    result = analyze_clusters(rows, labels)

    # Cluster 0 should have 2 members; label -1 excluded
    assert len(result) == 1
    assert result[0]["count"] == 2
    assert result[0]["cluster_id"] == 0


def test_analyze_clusters_structure():
    """Each cluster dict must contain required keys with correct types."""
    rows = [
        {"option_name": "Fan Kit",   "vendor": "hpe", "source_file": "f.xlsx"},
        {"option_name": "Fan Module","vendor": "dell","source_file": "g.xlsx"},
    ]
    labels = [1, 1]
    result = analyze_clusters(rows, labels)

    assert len(result) == 1
    c = result[0]
    assert "cluster_id" in c
    assert "count" in c
    assert isinstance(c["vendors"], list)
    assert isinstance(c["top_terms"], list)
    assert isinstance(c["examples"], list)
    assert c["count"] == 2
    assert set(c["vendors"]) == {"hpe", "dell"}


# ---------------------------------------------------------------------------
# heuristic_mapping
# ---------------------------------------------------------------------------

def test_heuristic_mapping_known_keyword():
    clusters = [{
        "cluster_id": 0,
        "count": 4,
        "vendors": ["hpe"],
        "source_files": ["f.xlsx"],
        "examples": ["Fan Kit for DL380"],
        "top_terms": ["fan", "kit"],
    }]
    result = heuristic_mapping(clusters)
    assert result[0]["proposed_device_type"] == "fan"
    assert result[0]["confidence"] == "heuristic"


def test_heuristic_mapping_unknown_returns_manual_review():
    clusters = [{
        "cluster_id": 1,
        "count": 3,
        "vendors": ["cisco"],
        "source_files": ["f.xlsx"],
        "examples": ["Zephyr Quantum Module"],
        "top_terms": ["zephyr", "quantum", "module"],
    }]
    result = heuristic_mapping(clusters)
    assert result[0]["proposed_device_type"] == ""
    assert result[0]["confidence"] == "manual_review"
    assert result[0]["suggested_yaml_rule"] == ""


# ---------------------------------------------------------------------------
# write_cluster_summary
# ---------------------------------------------------------------------------

def test_write_cluster_summary_creates_xlsx(tmp_path):
    clusters = [{
        "cluster_id": 0,
        "count": 5,
        "vendors": ["hpe"],
        "source_files": ["f.xlsx"],
        "examples": ["Fan Kit", "Fan Module", "Fan Assembly"],
        "top_terms": ["fan"],
        "proposed_device_type": "fan",
        "confidence": "heuristic",
        "suggested_yaml_rule": "fan",
    }]
    write_cluster_summary(clusters, tmp_path, min_cluster_size=3)
    assert (tmp_path / "cluster_summary.xlsx").exists()


def test_write_cluster_summary_updates_json(tmp_path):
    """Existing audit_report.json should gain a 'clusters' key."""
    report_path = tmp_path / "audit_report.json"
    report_path.write_text(json.dumps({"meta": {}}), encoding="utf-8")

    clusters = [{
        "cluster_id": 0,
        "count": 4,
        "vendors": ["dell"],
        "source_files": ["d.xlsx"],
        "examples": ["CPU Xeon"],
        "top_terms": ["cpu"],
        "proposed_device_type": "cpu",
        "confidence": "heuristic",
        "suggested_yaml_rule": "cpu",
    }]
    write_cluster_summary(clusters, tmp_path, min_cluster_size=3)

    report = json.loads(report_path.read_text(encoding="utf-8"))
    assert "clusters" in report
    assert report["clusters"]["total_clusters"] == 1


def test_write_cluster_summary_filters_small_clusters(tmp_path):
    """Clusters below min_cluster_size must not appear in the output xlsx."""
    clusters = [
        {
            "cluster_id": 0, "count": 10, "vendors": ["hpe"], "source_files": ["a.xlsx"],
            "examples": ["Fan Kit"], "top_terms": ["fan"],
            "proposed_device_type": "fan", "confidence": "heuristic", "suggested_yaml_rule": "fan",
        },
        {
            "cluster_id": 1, "count": 1, "vendors": ["hpe"], "source_files": ["a.xlsx"],
            "examples": ["Odd Part"], "top_terms": ["odd"],
            "proposed_device_type": "", "confidence": "manual_review", "suggested_yaml_rule": "",
        },
    ]
    write_cluster_summary(clusters, tmp_path, min_cluster_size=3)

    import openpyxl
    wb = openpyxl.load_workbook(tmp_path / "cluster_summary.xlsx")
    ws = wb.active
    # Only header row + 1 data row (cluster 0 with count=10)
    assert ws.max_row == 2


# ---------------------------------------------------------------------------
# print_dry_run_report
# ---------------------------------------------------------------------------

def test_print_dry_run_report_counts_by_vendor(capsys):
    candidates = [
        {"vendor": "hpe"},
        {"vendor": "hpe"},
        {"vendor": "dell"},
    ]
    print_dry_run_report(candidates)
    captured = capsys.readouterr().out
    assert "3" in captured       # total count
    assert "hpe" in captured
    assert "dell" in captured


# ---------------------------------------------------------------------------
# build_parser — CLI contract
# ---------------------------------------------------------------------------

def test_build_parser_output_dir_required():
    parser = build_parser()
    with pytest.raises(SystemExit):
        parser.parse_args([])   # --output-dir missing → error exit


def test_build_parser_defaults():
    parser = build_parser()
    args = parser.parse_args(["--output-dir", "/tmp/out"])
    assert args.min_cluster_size == 3
    assert args.max_clusters == 50
    assert args.vendor is None
    assert args.dry_run is False


# ---------------------------------------------------------------------------
# load_candidate_rows
# ---------------------------------------------------------------------------

def test_load_candidate_rows_empty_dir(tmp_path):
    """Empty directory returns empty list."""
    assert load_candidate_rows(tmp_path) == []


def test_load_candidate_rows_audited_E2(tmp_path):
    """Audited xlsx with E2 pipeline_check → exactly one row."""
    run_dir = tmp_path / "dell_run" / "run-2026-01-01__00-00-00-test"
    run_dir.mkdir(parents=True)

    path = run_dir / "test_annotated_audited.xlsx"
    _make_xlsx(
        path,
        ["option_name", "entity_type", "device_type", "hw_type", "pipeline_check"],
        [["Some Widget", "UNKNOWN", "", "", "E2: UNKNOWN entity_type"]],
    )

    rows = load_candidate_rows(tmp_path)

    assert len(rows) == 1
    assert rows[0]["option_name"] == "Some Widget"
    assert rows[0]["vendor"] == "dell"


def test_load_candidate_rows_annotated_unknown(tmp_path):
    """Annotated xlsx (no pipeline_check) with UNKNOWN entity → included."""
    run_dir = tmp_path / "dell_run" / "run-2026-01-01__00-00-00-test"
    run_dir.mkdir(parents=True)

    path = run_dir / "test_annotated.xlsx"
    _make_xlsx(
        path,
        ["option_name", "entity_type", "device_type", "hw_type"],
        [["Mystery Part", "UNKNOWN", "", ""]],
    )

    rows = load_candidate_rows(tmp_path)

    assert len(rows) == 1
    assert rows[0]["entity_type"] == "UNKNOWN"
    assert rows[0]["option_name"] == "Mystery Part"


def test_load_candidate_rows_filters_ok(tmp_path):
    """Audited xlsx with OK pipeline_check → zero rows."""
    run_dir = tmp_path / "dell_run" / "run-2026-01-01__00-00-00-test"
    run_dir.mkdir(parents=True)

    path = run_dir / "test_annotated_audited.xlsx"
    _make_xlsx(
        path,
        ["option_name", "entity_type", "device_type", "hw_type", "pipeline_check"],
        [["Good Part", "HW", "cpu", "cpu", "OK"]],
    )

    rows = load_candidate_rows(tmp_path)

    assert rows == []


def test_load_candidate_rows_vendor_filter(tmp_path):
    """Vendor filter returns only matching vendor."""
    for vendor in ("dell", "hpe"):
        run_dir = tmp_path / f"{vendor}_run" / f"run-2026-01-01__00-00-00-{vendor}"
        run_dir.mkdir(parents=True)

        path = run_dir / f"{vendor}_test_annotated.xlsx"
        _make_xlsx(
            path,
            ["option_name", "entity_type", "device_type", "hw_type"],
            [[f"{vendor} Part", "UNKNOWN", "", ""]],
        )

    rows = load_candidate_rows(tmp_path, vendor_filter="hpe")

    assert len(rows) == 1
    assert rows[0]["vendor"] == "hpe"
    assert "hpe" in rows[0]["option_name"].lower()


def test_load_candidate_rows_hw_no_types(tmp_path):
    """HW row without device_type and hw_type → included."""
    run_dir = tmp_path / "dell_run" / "run-2026-01-01__00-00-00-test"
    run_dir.mkdir(parents=True)

    path = run_dir / "test_annotated.xlsx"
    _make_xlsx(
        path,
        ["option_name", "entity_type", "device_type", "hw_type"],
        [["Bare HW Item", "HW", "", ""]],
    )

    rows = load_candidate_rows(tmp_path)

    assert len(rows) == 1
    assert rows[0]["entity_type"] == "HW"
    assert rows[0]["device_type"] == ""
    assert rows[0]["hw_type"] == ""


# ---------------------------------------------------------------------------
# _load_xlsx — robustness
# ---------------------------------------------------------------------------

def test_load_xlsx_no_known_columns(tmp_path):
    """XLSX with no known columns → fallback header=0, returns valid DataFrame."""
    path = tmp_path / "unknown_cols.xlsx"
    _make_xlsx(path, ["A", "B", "C"], [["1", "2", "3"]])
    df = _load_xlsx(path)
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 1
    assert len(df.columns) == 3


def test_load_xlsx_with_known_columns(tmp_path):
    """XLSX with known columns → auto-detected header, columns accessible."""
    path = tmp_path / "known_cols.xlsx"
    _make_xlsx(path, ["entity_type", "option_name", "device_type"],
               [["HW", "CPU Xeon", "cpu"]])
    df = _load_xlsx(path)
    assert isinstance(df, pd.DataFrame)
    cols_lower = {c.strip().lower() for c in df.columns}
    assert "entity_type" in cols_lower
    assert len(df) == 1


def test_load_xlsx_corrupt_file(tmp_path):
    """Corrupt file → raises any Exception (caught by load_candidate_rows)."""
    path = tmp_path / "bad.xlsx"
    path.write_bytes(b"this is not a valid xlsx file at all")
    with pytest.raises(Exception):
        _load_xlsx(path)


def test_load_xlsx_empty_data_rows(tmp_path):
    """XLSX with header but zero data rows → returns DataFrame with 0 rows."""
    path = tmp_path / "empty_data.xlsx"
    _make_xlsx(path, ["entity_type", "option_name"], [])
    df = _load_xlsx(path)
    assert isinstance(df, pd.DataFrame)
    assert len(df) == 0
