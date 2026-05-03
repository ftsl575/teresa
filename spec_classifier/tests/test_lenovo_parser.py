"""Tests for Lenovo DCSC parser (no external input files)."""

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from src.vendors.lenovo.adapter import LenovoAdapter
from src.vendors.lenovo.parser import parse_excel


_DEFAULT_HEADER = [
    "Part number",
    None,
    "Product Description",
    None,
    "Qty",
    "Price",
    "Total Part Price",
    "Export Control",
]


def _make_lenovo_xlsx(
    path,
    data_rows,
    *,
    include_terms=True,
    sheet_title="Quote",
    extra_rows_before_header=None,
    header_cells=None,
):
    """Create a minimal Lenovo DCSC Quote-like xlsx for testing."""
    extra_rows_before_header = extra_rows_before_header or []
    header_cells = list(header_cells) if header_cells is not None else list(_DEFAULT_HEADER)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_title

    # Row 0 (1-based: row 1)
    ws.append([None, None, "Data Center Solution Configurator Quote", None, None])
    ws.append([None, None, "Prepared for:", "Prepared by:"])
    ws.append([None, None, None, "Price Date:", "26-Feb-26"])
    ws.append([])
    ws.append([None, None, "Your final configuration..."])
    for row in extra_rows_before_header:
        ws.append(list(row))
    ws.append(header_cells)
    # Sub-header row (Part number empty — skipped until first real BOM line)
    ws.append([None, None, None, None, None, "(per unit)\nUS Dollar"])
    ws.append([])

    for row in data_rows:
        ws.append(row)

    if include_terms:
        ws.append([])
        ws.append(["TERMS AND CONDITIONS:", None, "Some legal text..."])

    wb.save(path)
    wb.close()


def test_header_row_index(tmp_path):
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(p, [
        ["BYW4", None, "Intel Xeon Silver 4510", None, 2, 100.0, None, None],
    ])
    rows, hdr = parse_excel(str(p))
    assert hdr == 5


def test_data_start_at_row_8(tmp_path):
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(p, [
        ["BYW4", None, "Intel Xeon Silver 4510", None, 2, 100.0, None, None],
    ], include_terms=False)
    rows, _ = parse_excel(str(p))
    assert len(rows) == 1
    assert rows[0]["Part number"] == "BYW4"
    assert rows[0]["Product Description"] == "Intel Xeon Silver 4510"
    assert rows[0]["Qty"] == 2
    assert rows[0]["Price"] == 100.0
    assert rows[0]["__row_index__"] == 9  # 0-based row 8 → 1-based row 9


def test_terms_and_conditions_stops_parsing(tmp_path):
    """Empty row before T&C is included; T&C itself is excluded."""
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(p, [
        ["BYW4", None, "CPU", None, 2, 100.0, None, None],
        ["BWHS", None, "RAM", None, 8, 50.0, None, None],
    ], include_terms=True)
    rows, _ = parse_excel(str(p))
    # 2 data rows + 1 empty separator before T&C
    assert len(rows) == 3
    assert rows[0]["Part number"] == "BYW4"
    assert rows[1]["Part number"] == "BWHS"
    assert rows[2]["Part number"] is None  # empty separator


def test_multi_server_cto_rows(tmp_path):
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(p, [
        ["7D73CTO1WW", None, "Server : ThinkSystem SR630 V3", None, 1, 5000.0, None, None],
        ["BYW4", None, "Intel Xeon Silver 4510", None, 2, 100.0, None, None],
        [None, None, None, None, None, None, None, None],  # separator
        ["7D9ACTO1WW", None, "Server 1C DB : ThinkSystem SR665 V3", None, 1, 8000.0, None, None],
        ["BWHS", None, "RAM Module", None, 8, 50.0, None, None],
    ], include_terms=False)
    rows, _ = parse_excel(str(p))
    assert len(rows) == 5
    assert rows[0]["Part number"] == "7D73CTO1WW"
    assert rows[3]["Part number"] == "7D9ACTO1WW"


def test_missing_sheet_raises(tmp_path):
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SomeOther"
    ws.append(["data"])
    wb.save(p)
    wb.close()
    with pytest.raises(ValueError, match="No Lenovo DCSC header") as exc:
        parse_excel(str(p))
    assert "SomeOther" in str(exc.value)


def test_header_row_found_when_shifted_down(tmp_path):
    """Extra preamble rows move the marker header below legacy row 5."""
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(
        p,
        [["BYW4", None, "Intel Xeon Silver 4510", None, 2, 100.0, None, None]],
        include_terms=False,
        extra_rows_before_header=[[]],
    )
    rows, hdr = parse_excel(str(p))
    assert hdr == 6
    assert len(rows) == 1
    assert rows[0]["Part number"] == "BYW4"
    assert rows[0]["__row_index__"] == 10


def test_renamed_sheet_fallback(tmp_path):
    """Parser discovers BOM / renamed sheet via header scan."""
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(
        p,
        [["BYW4", None, "CPU", None, 2, 100.0, None, None]],
        include_terms=False,
        sheet_title="BOM",
    )
    assert LenovoAdapter().can_parse(str(p))
    rows, hdr = parse_excel(str(p))
    assert hdr >= 5
    assert rows[0]["Part number"] == "BYW4"


def test_discount_column_inserted_before_price(tmp_path):
    header = [
        "Part number",
        None,
        "Product Description",
        None,
        "Qty",
        "Discount",
        "Price",
        "Total Part Price",
        "Export Control",
    ]
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(
        p,
        [["BYW4", None, "CPU line", None, 2, 5.0, 100.0, None, None]],
        include_terms=False,
        header_cells=header,
    )
    rows, _ = parse_excel(str(p))
    assert rows[0]["Qty"] == 2
    assert rows[0]["Price"] == 100.0


def test_permuted_columns(tmp_path):
    header = [
        "Product Description",
        None,
        "Part number",
        None,
        "Export Control",
        "Qty",
        "Price",
        "Total Part Price",
    ]
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(
        p,
        [["Intel Silver", None, "BYW4", None, "No", 3, 99.5, None]],
        include_terms=False,
        header_cells=header,
    )
    rows, _ = parse_excel(str(p))
    assert rows[0]["Part number"] == "BYW4"
    assert rows[0]["Product Description"] == "Intel Silver"
    assert rows[0]["Qty"] == 3
    assert rows[0]["Price"] == 99.5
    assert rows[0]["Export Control"] == "No"


def test_regression_internal_empty_row_still_separator_after_refactor(tmp_path):
    """Ensure pre-data blank skipping does not drop blank rows between BOM lines."""
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(
        p,
        [
            ["BYW4", None, "CPU", None, 2, 100.0, None, None],
            [None, None, None, None, None, None, None, None],
            ["BWHS", None, "RAM", None, 8, 50.0, None, None],
        ],
        include_terms=False,
    )
    rows, _ = parse_excel(str(p))
    assert len(rows) == 3
    assert rows[1]["Part number"] is None


def test_empty_rows_included_as_separators(tmp_path):
    p = tmp_path / "test.xlsx"
    _make_lenovo_xlsx(p, [
        ["BYW4", None, "CPU", None, 2, 100.0, None, None],
        [None, None, None, None, None, None, None, None],
        ["BWHS", None, "RAM", None, 8, 50.0, None, None],
    ], include_terms=False)
    rows, _ = parse_excel(str(p))
    assert len(rows) == 3
    assert rows[1]["Part number"] is None
    assert rows[1]["Product Description"] is None


def test_file_not_found():
    with pytest.raises(FileNotFoundError):
        parse_excel("/nonexistent/path/file.xlsx")


def test_adapter_get_source_sheet_name_reflects_actual_sheet(tmp_path):
    """get_source_sheet_name() returns the sheet parse() actually used."""
    # Default Quote sheet
    p1 = tmp_path / "quote.xlsx"
    _make_lenovo_xlsx(p1, [["BYW4", None, "CPU", None, 1, 10.0, None, None]],
                      include_terms=False)
    a1 = LenovoAdapter()
    assert a1.get_source_sheet_name() is None  # before parse()
    a1.parse(str(p1))
    assert a1.get_source_sheet_name() == "Quote"

    # Renamed sheet
    p2 = tmp_path / "bom.xlsx"
    _make_lenovo_xlsx(p2, [["BYW4", None, "CPU", None, 1, 10.0, None, None]],
                      include_terms=False, sheet_title="BOM")
    a2 = LenovoAdapter()
    a2.parse(str(p2))
    assert a2.get_source_sheet_name() == "BOM"

    # Quote w availability fallback (no plain Quote sheet)
    p3 = tmp_path / "qwa.xlsx"
    _make_lenovo_xlsx(p3, [["BYW4", None, "CPU", None, 1, 10.0, None, None]],
                      include_terms=False, sheet_title="Quote w availability")
    a3 = LenovoAdapter()
    a3.parse(str(p3))
    assert a3.get_source_sheet_name() == "Quote w availability"


def test_get_source_sheet_name_resets_on_parse_error(tmp_path):
    """Failed parse clears cached sheet name so it does not leak from prior file."""
    p_good = tmp_path / "good.xlsx"
    _make_lenovo_xlsx(p_good, [["BYW4", None, "CPU", None, 1, 10.0, None, None]],
                      include_terms=False)
    p_bad = tmp_path / "bad.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "SomeOther"
    ws.append(["data"])
    wb.save(p_bad)
    wb.close()

    adapter = LenovoAdapter()
    adapter.parse(str(p_good))
    assert adapter.get_source_sheet_name() == "Quote"

    with pytest.raises(ValueError, match="No Lenovo DCSC header"):
        adapter.parse(str(p_bad))
    assert adapter.get_source_sheet_name() is None
