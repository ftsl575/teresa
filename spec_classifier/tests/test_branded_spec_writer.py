"""
Tests for branded_spec_writer: file creation, absent block, title, multiple BASE blocks, UNKNOWN.
Build NormalizedRow and ClassificationResult manually; do not run pipeline.
"""

import pytest
import openpyxl
from pathlib import Path

from src.core.normalizer import NormalizedRow, RowKind
from src.core.classifier import ClassificationResult, EntityType
from src.core.state_detector import State
from src.outputs.branded_spec_writer import generate_branded_spec


def _nrow(idx: int, option_name: str, skus=None, qty=1, module_name="", **kwargs) -> NormalizedRow:
    """Build NormalizedRow with sensible defaults."""
    return NormalizedRow(
        source_row_index=idx,
        row_kind=RowKind.ITEM,
        group_name=kwargs.get("group_name"),
        group_id=kwargs.get("group_id"),
        product_name=kwargs.get("product_name"),
        module_name=module_name,
        option_name=option_name,
        option_id=kwargs.get("option_id"),
        skus=skus or [],
        qty=qty,
        option_price=kwargs.get("option_price", 0.0),
    )


def _all_cell_values(ws) -> set:
    """Collect all non-None cell values as strings for content checks."""
    out = set()
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                out.add(str(cell.value).strip())
    return out


# ── Test 1: file created, openpyxl opens, max_row > 0 ───────────────────────
def test_branded_spec_creates_file(tmp_path):
    """1 BASE + 2 HW PRESENT → file exists, openpyxl opens, ws.max_row > 0."""
    normalized_rows = [
        _nrow(1, "Server", skus=["SKU-BASE"], qty=1),
        _nrow(2, "HW One", skus=["SKU-1"], qty=1),
        _nrow(3, "HW Two", skus=["SKU-2"], qty=2),
    ]
    classification_results = [
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.PRESENT, "HW-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.PRESENT, "HW-002"),
    ]
    out_path = tmp_path / "branded.xlsx"
    result = generate_branded_spec(
        normalized_rows=normalized_rows,
        classification_results=classification_results,
        source_filename="test.xlsx",
        output_path=out_path,
    )
    assert result == out_path
    assert out_path.exists()
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    assert ws.max_row > 0
    wb.close()


# ── Test 2: absent block visible when SHOW_ABSENT_BLOCK = True ────────────────
def test_branded_spec_absent_block_present(tmp_path):
    """1 BASE + 1 HW PRESENT + 1 HW ABSENT, SHOW_ABSENT_BLOCK=True → 'Не установлено / Не выбрано' in file."""
    normalized_rows = [
        _nrow(1, "Server", skus=["SKU-BASE"], qty=1),
        _nrow(2, "HW Present", skus=["SKU-P"], qty=1),
        _nrow(3, "HW Absent", skus=["SKU-A"], qty=0),
    ]
    classification_results = [
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.PRESENT, "HW-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.ABSENT, "HW-002"),
    ]
    out_path = tmp_path / "branded.xlsx"
    generate_branded_spec(
        normalized_rows=normalized_rows,
        classification_results=classification_results,
        source_filename="test.xlsx",
        output_path=out_path,
    )
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    values = _all_cell_values(ws)
    wb.close()
    assert "Не установлено / Не выбрано" in values


# ── Test 3: absent block hidden when SHOW_ABSENT_BLOCK = False ──────────────
def test_branded_spec_absent_block_hidden(tmp_path):
    """Same as test 2; set SHOW_ABSENT_BLOCK=False, then restore. 'Не установлено / Не выбрано' must NOT appear."""
    import src.outputs.branded_spec_writer as bsw
    normalized_rows = [
        _nrow(1, "Server", skus=["SKU-BASE"], qty=1),
        _nrow(2, "HW Present", skus=["SKU-P"], qty=1),
        _nrow(3, "HW Absent", skus=["SKU-A"], qty=0),
    ]
    classification_results = [
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.PRESENT, "HW-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.ABSENT, "HW-002"),
    ]
    out_path = tmp_path / "branded.xlsx"
    try:
        bsw.SHOW_ABSENT_BLOCK = False
        generate_branded_spec(
            normalized_rows=normalized_rows,
            classification_results=classification_results,
            source_filename="test.xlsx",
            output_path=out_path,
        )
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        values = _all_cell_values(ws)
        wb.close()
        assert "Не установлено / Не выбрано" not in values
    finally:
        bsw.SHOW_ABSENT_BLOCK = True


# ── Test 4: only PRESENT in sections when absent hidden; DISABLED excluded ───
def test_branded_spec_only_present_in_sections(tmp_path):
    """1 BASE + 1 HW PRESENT (PRESENT_ITEM) + 1 HW ABSENT + 1 HW DISABLED, SHOW_ABSENT_BLOCK=False.
    PRESENT_ITEM in file; DISABLED option_name not in file."""
    import src.outputs.branded_spec_writer as bsw
    normalized_rows = [
        _nrow(1, "Server", skus=["SKU-BASE"], qty=1),
        _nrow(2, "PRESENT_ITEM", skus=["SKU-P"], qty=1),
        _nrow(3, "AbsentItem", skus=["SKU-A"], qty=0),
        _nrow(4, "DISABLED_ITEM", skus=["SKU-D"], qty=0),
    ]
    classification_results = [
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.PRESENT, "HW-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.ABSENT, "HW-002"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.DISABLED, "HW-003"),
    ]
    out_path = tmp_path / "branded.xlsx"
    try:
        bsw.SHOW_ABSENT_BLOCK = False
        generate_branded_spec(
            normalized_rows=normalized_rows,
            classification_results=classification_results,
            source_filename="test.xlsx",
            output_path=out_path,
        )
        wb = openpyxl.load_workbook(out_path)
        ws = wb.active
        values = _all_cell_values(ws)
        wb.close()
        assert "PRESENT_ITEM" in values
        assert "DISABLED_ITEM" not in values
    finally:
        bsw.SHOW_ABSENT_BLOCK = True


# ── Test 5: title contains source name (stem) ────────────────────────────────
def test_branded_spec_title_contains_source_name(tmp_path):
    """1 BASE, source_filename='my_spec_test.xlsx' → some cell contains 'my_spec_test'."""
    normalized_rows = [_nrow(1, "Server", skus=["SKU-BASE"], qty=1)]
    classification_results = [
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-001"),
    ]
    out_path = tmp_path / "branded.xlsx"
    generate_branded_spec(
        normalized_rows=normalized_rows,
        classification_results=classification_results,
        source_filename="my_spec_test.xlsx",
        output_path=out_path,
    )
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    values = _all_cell_values(ws)
    wb.close()
    found = any("my_spec_test" in v for v in values)
    assert found, "Expected some cell to contain 'my_spec_test'"


# ── Test 6: multiple BASE blocks ────────────────────────────────────────────
def test_branded_spec_multiple_base_blocks(tmp_path):
    """2 BASE (Server_A, Server_B), each with 1 HW PRESENT → Server_A and Server_B in file."""
    normalized_rows = [
        _nrow(1, "Server_A", skus=["SKU-A"], qty=1),
        _nrow(2, "HW A1", skus=["SKU-A1"], qty=1),
        _nrow(3, "Server_B", skus=["SKU-B"], qty=1),
        _nrow(4, "HW B1", skus=["SKU-B1"], qty=1),
    ]
    classification_results = [
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-001"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.PRESENT, "HW-001"),
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-002"),
        ClassificationResult(RowKind.ITEM, EntityType.HW, State.PRESENT, "HW-002"),
    ]
    out_path = tmp_path / "branded.xlsx"
    generate_branded_spec(
        normalized_rows=normalized_rows,
        classification_results=classification_results,
        source_filename="test.xlsx",
        output_path=out_path,
    )
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    values = _all_cell_values(ws)
    wb.close()
    assert "Server_A" in values
    assert "Server_B" in values


# ── Test 7: UNKNOWN entity type always included ─────────────────────────────
def test_branded_spec_unknown_entity_type(tmp_path):
    """1 BASE + 1 UNKNOWN (state=None). option_name of UNKNOWN row must be in file."""
    normalized_rows = [
        _nrow(1, "Server", skus=["SKU-BASE"], qty=1),
        _nrow(2, "UnknownOption", skus=["SKU-U"], qty=1),
    ]
    classification_results = [
        ClassificationResult(RowKind.ITEM, EntityType.BASE, State.PRESENT, "BASE-001"),
        ClassificationResult(RowKind.ITEM, EntityType.UNKNOWN, None, "UNKNOWN-000"),
    ]
    out_path = tmp_path / "branded.xlsx"
    generate_branded_spec(
        normalized_rows=normalized_rows,
        classification_results=classification_results,
        source_filename="test.xlsx",
        output_path=out_path,
    )
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    values = _all_cell_values(ws)
    wb.close()
    assert "UnknownOption" in values
