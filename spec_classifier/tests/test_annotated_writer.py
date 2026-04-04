"""
Smoke test for annotated source Excel export.
Uses robust header detection so annotated files with a preamble (e.g. Solution Info) are parsed correctly.
"""

import pytest
import pandas as pd
import openpyxl
from pathlib import Path

from main import _get_adapter
from src.rules.rules_engine import RuleSet
from src.core.classifier import classify_row, ClassificationResult, EntityType
from src.core.normalizer import NormalizedRow, RowKind
from src.core.state_detector import State
from src.outputs.annotated_writer import generate_annotated_source_excel

from conftest import project_root, get_input_root_dell
from tests.helpers import find_annotated_header_row, read_annotated_excel


def test_annotated_writer_reads_named_sheet(tmp_path):
    """sheet_name='BOM' reads from the BOM sheet, not sheet index 0."""
    # Build a two-sheet workbook:
    #   Sheet1 (index 0) — 5 rows of junk data (must NOT be read)
    #   BOM              — 1 header row + 1 data row (SHOULD be read)
    wb = openpyxl.Workbook()
    ws0 = wb.active
    ws0.title = "Sheet1"
    for i in range(5):
        ws0.append([f"sheet0_row_{i}", "dummy", "data"])

    ws_bom = wb.create_sheet("BOM")
    ws_bom.append(["Product #", "Product Description", "Qty"])   # row 1 — header
    ws_bom.append(["SKU-CPU-001", "Intel Xeon Gold 5416S", "1"]) # row 2 — data

    src_path = tmp_path / "hp_test.xlsx"
    wb.save(str(src_path))

    # Minimal pipeline objects: one ITEM row at source_row_index=2 (BOM row 2)
    norm = NormalizedRow(
        source_row_index=2,
        row_kind=RowKind.ITEM,
        group_name=None, group_id=None, product_name=None,
        module_name="", option_name="Intel Xeon Gold 5416S",
        option_id="SKU-CPU-001", skus=["SKU-CPU-001"], qty=1, option_price=0.0,
    )
    result = ClassificationResult(
        row_kind=RowKind.ITEM,
        entity_type=EntityType.HW,
        state=State.PRESENT,
        matched_rule_id="HW-TEST-001",
        device_type="cpu",
        hw_type="cpu",
    )

    out_path = generate_annotated_source_excel(
        raw_rows=[],
        normalized_rows=[norm],
        classification_results=[result],
        original_excel_path=src_path,
        run_folder=tmp_path,
        header_row_index=0,
        sheet_name="BOM",
        extra_cols=[],
    )

    assert out_path.exists(), "annotated workbook should be created"
    assert out_path.name == "hp_test_annotated.xlsx"

    df_out = pd.read_excel(out_path, header=None, engine="openpyxl")
    # BOM has 2 rows; Sheet1 has 5 rows — wrong sheet read → wrong row count
    assert len(df_out) == 2, (
        f"Expected 2 rows (from BOM sheet), got {len(df_out)} "
        f"(Sheet1 has 5 rows — reading wrong sheet if 5)"
    )
    # Header row (row 0) must contain the annotation column labels
    header_values = set(str(v) for v in df_out.iloc[0].dropna())
    assert "Entity Type" in header_values
    assert "hw_type" in header_values
    # Data row (row 1) must have the classification value
    data_values = set(str(v) for v in df_out.iloc[1].dropna())
    assert "HW" in data_values


def test_annotated_excel_exists_same_rows_has_entity_type_state_and_item_values(tmp_path):
    """Run pipeline on dl1.xlsx, generate annotated Excel; assert file exists, row count matches, has Entity Type/State columns, ITEM rows filled."""
    root = project_root()
    input_path = get_input_root_dell() / "dl1.xlsx"
    if not input_path.exists():
        pytest.skip(f"Input not found: {input_path} (set paths.input_root in config.local.yaml)")

    adapter = _get_adapter("dell", {})
    raw_rows, header_row_index = adapter.parse(str(input_path))
    normalized_rows = adapter.normalize(raw_rows)
    ruleset = RuleSet.load(str(root / "rules" / "dell_rules.yaml"))
    classification_results = [classify_row(r, ruleset) for r in normalized_rows]

    out_path = generate_annotated_source_excel(
        raw_rows, normalized_rows, classification_results, input_path, tmp_path,
        header_row_index=header_row_index,
        extra_cols=[],
    )

    assert out_path.exists(), "annotated workbook should exist"
    assert out_path.name == f"{input_path.stem}_annotated.xlsx"

    df_orig = pd.read_excel(input_path, header=None, engine="openpyxl")
    header_row_index, df_ann = read_annotated_excel(out_path)

    # Row count: annotated has same total rows as original; df_ann is header + data, so data rows = total - header - 1
    expected_data_rows = len(df_orig) - header_row_index - 1
    assert len(df_ann) == expected_data_rows, (
        f"Data row count must match: expected {expected_data_rows}, got {len(df_ann)}"
    )
    for col in ("Entity Type", "State", "device_type", "hw_type", "matched_rule_id"):
        assert col in df_ann.columns, f"Annotated file must have column {col!r}"
    # At least one ITEM row has non-empty Entity Type (e.g. BASE, HW)
    non_empty = df_ann["Entity Type"].dropna().astype(str).str.strip()
    non_empty = non_empty[(non_empty != "") & (non_empty != "Entity Type")]
    assert len(non_empty) > 0, "At least one ITEM row should have Entity Type filled"


def test_annotated_matched_rule_id_column_populated(tmp_path):
    """matched_rule_id column is present and non-empty for at least one ITEM row."""
    root = project_root()
    input_path = get_input_root_dell() / "dl1.xlsx"
    if not input_path.exists():
        pytest.skip(f"Input not found: {input_path} (set paths.input_root in config.local.yaml)")

    adapter = _get_adapter("dell", {})
    raw_rows, header_row_index = adapter.parse(str(input_path))
    normalized_rows = adapter.normalize(raw_rows)
    ruleset = RuleSet.load(str(root / "rules" / "dell_rules.yaml"))
    classification_results = [classify_row(r, ruleset) for r in normalized_rows]

    out_path = generate_annotated_source_excel(
        raw_rows, normalized_rows, classification_results, input_path, tmp_path,
        header_row_index=header_row_index,
        extra_cols=[],
    )

    _, df_ann = read_annotated_excel(out_path)
    assert "matched_rule_id" in df_ann.columns, "matched_rule_id column must be present"

    # At least one ITEM row must have a non-empty matched_rule_id
    rule_ids = df_ann["matched_rule_id"].dropna().astype(str).str.strip()
    rule_ids = rule_ids[(rule_ids != "") & (rule_ids != "matched_rule_id")]
    assert len(rule_ids) > 0, "At least one ITEM row should have matched_rule_id filled"
    # Spot-check: all non-empty values look like rule IDs (no spaces, alphanumeric + dashes)
    import re
    for rid in rule_ids:
        assert re.match(r"^[\w\-]+$", rid), f"Unexpected matched_rule_id format: {rid!r}"


def test_annotated_header_row_detection_with_preamble(tmp_path):
    """Regression: annotated xlsx with preamble (e.g. dl4) — parser must find header and read by Option ID."""
    root = project_root()
    input_path = get_input_root_dell() / "dl4.xlsx"
    if not input_path.exists():
        pytest.skip(f"Input not found: {input_path} (set paths.input_root in config.local.yaml)")

    adapter = _get_adapter("dell", {})
    raw_rows, header_row_index = adapter.parse(str(input_path))
    normalized_rows = adapter.normalize(raw_rows)
    ruleset = RuleSet.load(str(root / "rules" / "dell_rules.yaml"))
    classification_results = [classify_row(r, ruleset) for r in normalized_rows]

    out_path = generate_annotated_source_excel(
        raw_rows, normalized_rows, classification_results, input_path, tmp_path,
        header_row_index=header_row_index,
        extra_cols=[],
    )

    header_row_index = find_annotated_header_row(out_path)
    assert header_row_index is not None, "Header row must be found in annotated file"
    _, df = read_annotated_excel(out_path)
    assert "Option ID" in df.columns, "Option ID column must be present"
    assert "Entity Type" in df.columns and "hw_type" in df.columns

    # Spot-check: read a row by Option ID (e.g. GBEZWO8) — no KeyError
    option_col = df["Option ID"].astype(str).str.strip()
    row = df[option_col == "GBEZWO8"]
    assert len(row) >= 1, "Expected at least one row with Option ID GBEZWO8"
    first = row.iloc[0]
    assert first["Entity Type"] in ("HW", "BASE", "CONFIG", "LOGISTIC", "NOTE", "SOFTWARE", "SERVICE"), (
        "Entity Type should be a known type"
    )


def test_extra_cols_added_to_output(tmp_path):
    """extra_cols tuples become real columns in the annotated Excel, header and data populated."""
    import types

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product #", "Description"])  # row 1: header
    ws.append(["SKU-001", "Some product"])   # row 2: data
    src = tmp_path / "vendor_test.xlsx"
    wb.save(str(src))

    # Use SimpleNamespace so we can attach arbitrary vendor attributes
    norm = types.SimpleNamespace(source_row_index=2, row_kind=RowKind.ITEM, product_type="Server")
    result = ClassificationResult(
        row_kind=RowKind.ITEM,
        entity_type=EntityType.HW,
        state=State.PRESENT,
        matched_rule_id="HW-001",
        device_type="server",
        hw_type="server",
    )

    out = generate_annotated_source_excel(
        raw_rows=[],
        normalized_rows=[norm],
        classification_results=[result],
        original_excel_path=src,
        run_folder=tmp_path,
        header_row_index=0,
        extra_cols=[("product_type", "product_type")],
    )

    df = pd.read_excel(out, header=None, engine="openpyxl")
    header_vals = {str(v) for v in df.iloc[0]}
    assert "product_type" in header_vals, "extra_col header 'product_type' must appear in row 0"
    data_vals = {str(v) for v in df.iloc[1]}
    assert "Server" in data_vals, "extra_col value 'Server' must appear in data row"


def test_empty_extra_cols_no_vendor_columns(tmp_path):
    """extra_cols=() produces no vendor-extension columns — not even empty ones."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Module", "Option"])
    ws.append(["Processor", "Xeon Gold"])
    src = tmp_path / "dell_noextra.xlsx"
    wb.save(str(src))

    norm = NormalizedRow(
        source_row_index=2,
        row_kind=RowKind.ITEM,
        group_name=None, group_id=None, product_name=None,
        module_name="Processor", option_name="Xeon Gold",
        option_id=None, skus=[], qty=1, option_price=0.0,
    )
    result = ClassificationResult(
        row_kind=RowKind.ITEM,
        entity_type=EntityType.HW,
        state=State.PRESENT,
        matched_rule_id="HW-002",
        device_type="cpu",
        hw_type="cpu",
    )

    out = generate_annotated_source_excel(
        raw_rows=[],
        normalized_rows=[norm],
        classification_results=[result],
        original_excel_path=src,
        run_folder=tmp_path,
        header_row_index=0,
        extra_cols=(),
    )

    df = pd.read_excel(out, header=None, engine="openpyxl")
    col_headers = {str(v) for v in df.iloc[0]}
    vendor_cols = {
        "line_number", "service_duration_months",
        "product_type", "extended_price", "lead_time",
        "config_name", "is_factory_integrated",
    }
    unexpected = vendor_cols & col_headers
    assert not unexpected, f"No vendor columns expected with extra_cols=(); found: {unexpected}"
