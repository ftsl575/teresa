import openpyxl
from src.vendors.base import VendorAdapter
from src.vendors.xfusion.parser import parse_excel
from src.vendors.xfusion.normalizer import normalize_xfusion_rows


class XFusionAdapter(VendorAdapter):
    """
    xFusion FusionServer eDeal adapter.
    can_parse — Phase 0 (mutually exclusive twin with HuaweiAdapter).
    parse / normalize / get_rules_file / extras — Phase 1.
    Real classification rules land in Phase 3 (rules/xfusion_rules.yaml currently scaffold).
    """

    def __init__(self, config: dict = None):
        self._config = config or {}

    def can_parse(self, path: str) -> bool:
        """
        Positive signature for xFusion FusionServer eDeal export.

        Two-branch discriminator (mutually exclusive with Huawei — see
        VENDOR_FORMAT_SPEC.md (xfusion) §3.2):
          Branch A (Excel Desktop, formula evaluated): cached R8C8 contains
            "USD" but neither "FOB" nor "HONG KONG" -> xFusion
            (QF_SYS_TRADETERMDESC1 is empty in xFusion eDeal templates).
          Branch B (Excel Online, formula unevaluated): cached R8C8 empty AND
            sidecar sheet "Main Equipment Statistic" is PRESENT -> xFusion.

        Note: data_only=True is required — the price-header cell stores a
        formula that resolves differently for Huawei vs xFusion based on the
        injected default for the QF_SYS_TRADETERMDESC1 named range.
        """
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if "AllInOne" not in wb.sheetnames:
                return False
            ws = wb["AllInOne"]
            rows = list(ws.iter_rows(min_row=1, max_row=9, values_only=True))
            if len(rows) < 9:
                return False
            r0, r8 = rows[0], rows[8]
            if len(r0) < 3 or r0[2] is None or str(r0[2]).strip() != "COL_SORTNO.0":
                return False
            r8c8 = str(r8[8]).strip() if len(r8) > 8 and r8[8] is not None else ""
            # Branch A: cached header is bare USD (xFusion convention)
            if r8c8 and "FOB" not in r8c8.upper() and "HONG KONG" not in r8c8.upper():
                return True
            # Branch B: hidden prices — sidecar sheet is xFusion-only
            if not r8c8 and "Main Equipment Statistic" in wb.sheetnames:
                return True
            return False
        finally:
            wb.close()

    def parse(self, filepath: str):
        return parse_excel(filepath)

    def normalize(self, raw_rows):
        return normalize_xfusion_rows(raw_rows)

    def get_rules_file(self):
        vendor_rules = self._config.get("vendor_rules", {})
        return vendor_rules.get("xfusion", "rules/xfusion_rules.yaml")

    def get_vendor_stats(self, normalized_rows: list) -> dict:
        # Phase 1 minimal stub — real implementation in Phase 2 per CC Q4
        # (server_configs_count, unique_models_count, spare_parts_groups_count).
        return {}

    def get_source_sheet_name(self) -> str | None:
        return "AllInOne"

    def get_extra_cols(self) -> list[tuple[str, str]]:
        return [
            ("position_no",     "position_no"),
            ("model",           "model"),
            ("unit_qty",        "unit_qty"),
            ("total_price",     "total_price"),
            ("lead_time_days",  "lead_time_days"),
        ]

    def generates_branded_spec(self) -> bool:
        return True
