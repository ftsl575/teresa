"""
xFusion FusionServer eDeal Excel parser.
Sheet 'AllInOne', header row 8 (0-based), data from row 9.

Index-driven (column-by-position) — robust to xf10-style blanked R8 price
headers per VENDOR_FORMAT_SPEC.md (xfusion) §6.6.

Columns by fixed index:
  col2 = Position No  (e.g. ""/"1"/"1.1"/"1.1.1")
  col3 = Part Number  (SKU on ITEM rows; sub-category label on HEADER)
  col4 = Model        (vendor part code on ITEM; model name on level-1 HEADER)
  col5 = Description  (free text)
  col6 = Unit Qty.
  col7 = Qty.
  col8 = Unit Price   (USD; formula → requires data_only=True)
  col9 = Total Price  (USD; formula → requires data_only=True)
  col10 = Production LT (Days)  ("Uncertain" / "-" / numeric — kept as str)

No stop sentinel — read until end of sheet, skipping fully-empty separator
rows (col2 AND col3 both empty).
"""

import openpyxl
from pathlib import Path
from typing import List, Tuple


_SHEET_NAME = "AllInOne"
_HEADER_ROW = 8       # 0-based; display headers at Excel row 9
_DATA_START_ROW = 9   # 0-based; data begins at Excel row 10


def _is_sku_shape(value) -> bool:
    """
    xFusion SKU shape: alphanumeric (with '-' / '_'), no whitespace,
    ≤ 14 chars, must contain at least one digit.

    Sub-category labels at level-3 (e.g. "Memory", "Base Configuration",
    "Hard Disk(with 2.5\" Front Panel)-NVMe") fail one of these checks
    (whitespace, length, parens, or no-digit) and are routed to the
    normalizer as HEADER rows.
    """
    if value is None:
        return False
    s = str(value).strip()
    if not s:
        return False
    if any(ch.isspace() for ch in s):
        return False
    if len(s) > 14:
        return False
    if not any(c.isdigit() for c in s):
        return False
    return all(c.isalnum() or c in "-_" for c in s)


def parse_excel(filepath: str) -> Tuple[List[dict], int]:
    """
    Parse xFusion eDeal AllInOne sheet → (rows, header_row_index=8).

    Each row dict carries fixed-index columns plus __row_index__ (1-based).
    Fully-empty separator rows (col2 AND col3 both empty) are skipped — they
    do NOT survive into the result list (normalizer therefore never has to
    decide whether to reset rollup on them).
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if _SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{_SHEET_NAME}' not found in {filepath}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[_SHEET_NAME]
        all_rows = list(ws.iter_rows(values_only=True))

        rows: List[dict] = []

        for i in range(_DATA_START_ROW, len(all_rows)):
            raw = all_rows[i]
            excel_row_1based = i + 1

            col2 = raw[2]  if len(raw) >  2 else None
            col3 = raw[3]  if len(raw) >  3 else None
            col4 = raw[4]  if len(raw) >  4 else None
            col5 = raw[5]  if len(raw) >  5 else None
            col6 = raw[6]  if len(raw) >  6 else None
            col7 = raw[7]  if len(raw) >  7 else None
            col8 = raw[8]  if len(raw) >  8 else None
            col9 = raw[9]  if len(raw) >  9 else None
            col10 = raw[10] if len(raw) > 10 else None

            col2_str = str(col2).strip() if col2 is not None else ""
            col3_str = str(col3).strip() if col3 is not None else ""

            # Drop fully-empty separator rows; level-0 site headers (col2 empty,
            # col3 non-empty) survive — that's exactly the rollup boundary.
            if not col2_str and not col3_str:
                continue

            row_dict = {
                "Position No":         col2,
                "Part Number":         col3,
                "Model":               col4,
                "Description":         col5,
                "Unit Qty.":           col6,
                "Qty.":                col7,
                "Unit Price":          col8,
                "Total Price":         col9,
                "production_lt_days":  col10 if col10 is not None else "",
                "__row_index__":       excel_row_1based,
            }
            rows.append(row_dict)

        return (rows, _HEADER_ROW)
    finally:
        wb.close()
