import openpyxl
from src.vendors.base import VendorAdapter
from src.core.parser import parse_excel
from src.core.normalizer import normalize_row


class DellAdapter(VendorAdapter):
    def __init__(self, config: dict = None):
        self._config = config or {}

    def can_parse(self, path: str) -> bool:
        """Positive signature: first sheet contains 'Module Name' in first 20 rows."""
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            if not wb.sheetnames:
                return False
            ws = wb[wb.sheetnames[0]]
            for row in ws.iter_rows(max_row=20, values_only=True):
                for val in row:
                    if val is not None and str(val).strip() == "Module Name":
                        return True
            return False
        finally:
            wb.close()

    def parse(self, filepath: str):
        return parse_excel(filepath)

    def normalize(self, raw_rows):
        return [normalize_row(r) for r in raw_rows]

    def get_rules_file(self):
        vendor_rules = self._config.get("vendor_rules", {})
        return vendor_rules.get("dell", "rules/dell_rules.yaml")

    def get_vendor_stats(self, normalized_rows: list) -> dict:
        return {}

    def generates_branded_spec(self) -> bool:
        return True
