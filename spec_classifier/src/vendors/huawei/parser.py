"""
Huawei eDeal Excel parser.
Sheet 'AllInOne', header row 8 (0-based), data from row 9.
Columns by fixed index: col3=Part Number (primary classifier),
col5=Product Description, col6=Unit Quantity, col7=Qty,
col8=Total Price, col9=Extended Price.
Stop sentinel: col3 == "EOM" → stop reading (do not include that row).
Optional header-lookup columns: 'Production LT (Days)', 'EOM', 'EOS'.
"""

import openpyxl
from pathlib import Path
from typing import List, Tuple


_SHEET_NAME = "AllInOne"
_HEADER_ROW = 8
_DATA_START_ROW = 9
_STOP_SENTINEL = "EOM"


def _find_col_index(header_row_values: tuple, name: str, partial: bool = False) -> int:
    """Return 0-based column index of name in header_row_values, or -1 if not found."""
    for idx, cell in enumerate(header_row_values):
        if cell is None:
            continue
        cell_str = str(cell).strip()
        if partial:
            if name.lower() in cell_str.lower():
                return idx
        else:
            if cell_str == name:
                return idx
    return -1


def parse_excel(filepath: str) -> Tuple[List[dict], int]:
    """
    Parse Huawei eDeal AllInOne sheet.

    Returns (rows, header_row_index=8).
    Each row dict contains fixed-index columns plus optional header-lookup
    columns (production_lt_days, eom, eos) and __row_index__ (1-based Excel row).
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        if _SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{_SHEET_NAME}' not found in {filepath}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[_SHEET_NAME]
        all_rows = list(ws.iter_rows(values_only=True))

        # Resolve optional header-lookup column indices
        header_values = all_rows[_HEADER_ROW] if len(all_rows) > _HEADER_ROW else ()
        lt_col_idx = _find_col_index(header_values, "Production LT", partial=True)
        eom_col_idx = _find_col_index(header_values, "EOM", partial=False)
        eos_col_idx = _find_col_index(header_values, "EOS", partial=False)

        rows: List[dict] = []

        for i in range(_DATA_START_ROW, len(all_rows)):
            raw = all_rows[i]
            excel_row_1based = i + 1

            col0 = raw[0] if len(raw) > 0 else None
            col1 = raw[1] if len(raw) > 1 else None
            col2 = raw[2] if len(raw) > 2 else None
            col3 = raw[3] if len(raw) > 3 else None
            col4 = raw[4] if len(raw) > 4 else None
            col5 = raw[5] if len(raw) > 5 else None
            col6 = raw[6] if len(raw) > 6 else None
            col7 = raw[7] if len(raw) > 7 else None
            col8 = raw[8] if len(raw) > 8 else None
            col9 = raw[9] if len(raw) > 9 else None

            col3_str = str(col3).strip() if col3 is not None else ""
            if col3_str == _STOP_SENTINEL:
                break

            production_lt = (
                raw[lt_col_idx] if lt_col_idx >= 0 and len(raw) > lt_col_idx else ""
            )
            eom_val = (
                raw[eom_col_idx] if eom_col_idx >= 0 and len(raw) > eom_col_idx else ""
            )
            eos_val = (
                raw[eos_col_idx] if eos_col_idx >= 0 and len(raw) > eos_col_idx else ""
            )

            row_dict = {
                "Position No":          col2,
                "Part Number":          col3,
                "Product Description":  col5,
                "Unit Quantity":        col6,
                "Qty":                  col7,
                "Total Price":          col8,
                "Extended Price":       col9,
                "_col0_metadata":       col0,
                "_col1_sort_no":        col1,
                "_col4_aux":            col4,
                "production_lt_days":   production_lt if production_lt is not None else "",
                "eom":                  eom_val if eom_val is not None else "",
                "eos":                  eos_val if eos_val is not None else "",
                "__row_index__":        excel_row_1based,
            }
            rows.append(row_dict)

        return (rows, _HEADER_ROW)
    finally:
        wb.close()
