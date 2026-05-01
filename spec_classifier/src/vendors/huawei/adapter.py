import openpyxl
from src.vendors.base import VendorAdapter
from src.vendors.huawei.parser import parse_excel
from src.vendors.huawei.normalizer import normalize_huawei_rows


class HuaweiAdapter(VendorAdapter):
    def __init__(self, config: dict = None):
        self._config = config or {}

    def can_parse(self, path: str) -> bool:
        """
        Tightened Huawei eDeal positive signature (mutually exclusive with xFusion).

        Two-branch discriminator:
          Branch A (Excel Desktop, formula evaluated): cached R8C8 contains
            "FOB" or "HONG KONG" -> Huawei (per QF_SYS_TRADETERMDESC1 default).
          Branch B (Excel Online, formula unevaluated): cached R8C8 empty AND
            sidecar sheet "Main Equipment Statistic" is ABSENT -> Huawei.

        See VENDOR_FORMAT_SPEC.md (xfusion) §3.2 for the discrimination rationale
        and §3.5 for the 15-fixture verification matrix (5 hu + 10 xf, disjoint).
        """
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if "AllInOne" not in wb.sheetnames:
                return False
            ws = wb["AllInOne"]
            rows = list(ws.iter_rows(min_row=1, max_row=9, values_only=True))
            if len(rows) < 9:
                return False
            r0, r8 = rows[0], rows[8]
            if len(r0) < 3 or r0[2] is None or str(r0[2]).strip() != "COL_SORTNO.0":
                return False
            r8c8 = str(r8[8]).strip() if len(r8) > 8 and r8[8] is not None else ""
            # Branch A: cached header carries Huawei's FOB Incoterm
            if r8c8 and ("FOB" in r8c8.upper() or "HONG KONG" in r8c8.upper()):
                return True
            # Branch B: hidden prices — absence of xFusion sidecar implies Huawei
            if not r8c8 and "Main Equipment Statistic" not in wb.sheetnames:
                return True
            return False
        finally:
            wb.close()

    def parse(self, filepath: str):
        return parse_excel(filepath)

    def normalize(self, raw_rows):
        return normalize_huawei_rows(raw_rows)

    def get_rules_file(self):
        vendor_rules = self._config.get("vendor_rules", {})
        return vendor_rules.get("huawei", "rules/huawei_rules.yaml")

    def get_vendor_stats(self, normalized_rows: list) -> dict:
        return {}

    def get_source_sheet_name(self) -> str | None:
        return "AllInOne"

    def get_extra_cols(self) -> list[tuple[str, str]]:
        return [
            ("position_no",    "position_no"),
            ("unit_qty",       "unit_qty"),
            ("total_price",    "total_price"),
            ("lead_time_days", "lead_time_days"),
            ("eom",            "eom"),
            ("eos",            "eos"),
        ]

    def generates_branded_spec(self) -> bool:
        return True
