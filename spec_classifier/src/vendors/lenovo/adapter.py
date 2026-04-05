import openpyxl
from src.vendors.base import VendorAdapter
from src.vendors.lenovo.parser import parse_excel
from src.vendors.lenovo.normalizer import normalize_lenovo_rows


class LenovoAdapter(VendorAdapter):
    def __init__(self, config: dict = None):
        self._config = config or {}

    def can_parse(self, path: str) -> bool:
        """
        Positive signature:
        1. Sheet "Quote" present in workbook.
        2. Cell row 0, col 2 contains "Data Center Solution Configurator".
        """
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            if "Quote" not in wb.sheetnames:
                return False
            ws = wb["Quote"]
            row0 = next(ws.iter_rows(max_row=1, values_only=True), ())
            if len(row0) < 3 or row0[2] is None:
                return False
            return "Data Center Solution Configurator" in str(row0[2])
        finally:
            wb.close()

    def parse(self, filepath: str):
        return parse_excel(filepath)

    def normalize(self, raw_rows):
        return normalize_lenovo_rows(raw_rows)

    def get_rules_file(self):
        vendor_rules = self._config.get("vendor_rules", {})
        return vendor_rules.get("lenovo", "rules/lenovo_rules.yaml")

    def get_vendor_stats(self, normalized_rows: list) -> dict:
        cto_count = sum(
            1 for r in normalized_rows
            if getattr(r, "option_id", None) and "CTO" in str(r.option_id).upper()
        )
        export_controlled = sum(
            1 for r in normalized_rows
            if getattr(r, "export_control", "") == "Yes"
        )
        return {
            "cto_count": cto_count,
            "export_controlled_count": export_controlled,
        }

    def get_source_sheet_name(self) -> str | None:
        return "Quote"

    def get_extra_cols(self) -> list[tuple[str, str]]:
        return [("export_control", "export_control")]

    def generates_branded_spec(self) -> bool:
        return False
