"""
Lenovo DCSC (Data Center Solution Configurator) Excel parser.

Sheet selection: try "Quote", then "Quote w availability", then other sheets
(in workbook order), skipping DCSC utility sheets. The first sheet whose first
30 rows contain the marker header row wins.

Header row: found by scanning for all of Part number, Product Description,
Qty, Price, Export Control (exact names, stripped). Columns are read via
col_map -- not fixed indices. "Total Part Price" may be present (formulas);
it is never emitted in row dicts.

data_only=False keeps formula objects in Total Part Price if that column is
read later; we do not expose it in output.
"""

import openpyxl
from pathlib import Path
from typing import Dict, List, Optional, Sequence, Tuple

_STOP_SENTINEL = "terms and conditions"

HEADER_MARKERS: Tuple[str, ...] = (
    "Part number",
    "Product Description",
    "Qty",
    "Price",
    "Export Control",
)

DCSC_EXCLUDED_SHEETS = frozenset(
    {
        "Power Report",
        "ConfigGroupView",
        "Summary",
        "Message History",
    }
)


def ordered_sheet_candidates(sheetnames: Sequence[str]) -> List[str]:
    """Quote first, then Quote w availability, then other non-excluded sheets in workbook order."""
    out: List[str] = []
    seen: set = set()
    if "Quote" in sheetnames:
        out.append("Quote")
        seen.add("Quote")
    if "Quote w availability" in sheetnames:
        out.append("Quote w availability")
        seen.add("Quote w availability")
    for n in sheetnames:
        if n in DCSC_EXCLUDED_SHEETS or n in seen:
            continue
        out.append(n)
        seen.add(n)
    return out


def _header_scan_limit() -> int:
    return 30


def find_lenovo_header_in_rows(
    rows: Sequence[Tuple],
) -> Optional[Tuple[int, Dict[str, int]]]:
    """
    Scan rows (typically first 30) for the Lenovo DCSC marker header.
    Returns (0-based header row index, col_map) or None.
    """
    limit = min(len(rows), _header_scan_limit())
    for i in range(limit):
        row = rows[i]
        col_map: Dict[str, int] = {}
        for idx, val in enumerate(row):
            if val is not None:
                col_map[str(val).strip()] = idx
        if all(m in col_map for m in HEADER_MARKERS):
            return (i, col_map)
    return None


def workbook_has_lenovo_dcsc_header(filepath: str) -> bool:
    """True if any non-excluded sheet has the marker header in its first 30 rows."""
    path = Path(filepath)
    if not path.exists():
        return False
    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        for name in ordered_sheet_candidates(wb.sheetnames):
            ws = wb[name]
            chunk: List[Tuple] = []
            for j, row in enumerate(ws.iter_rows(values_only=True)):
                chunk.append(tuple(row))
                if j >= _header_scan_limit() - 1:
                    break
            if find_lenovo_header_in_rows(chunk) is not None:
                return True
        return False
    finally:
        wb.close()


def _first_nonempty_stripped(raw_row: Sequence) -> str:
    for v in raw_row:
        if v is None:
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def _cell_at(raw_row: Sequence, idx: int):
    if idx < len(raw_row):
        return raw_row[idx]
    return None


def _part_number_nonempty(raw_row: Sequence, pn_idx: int) -> bool:
    v = _cell_at(raw_row, pn_idx)
    if v is None:
        return False
    return bool(str(v).strip())


def parse_excel_with_sheet(filepath: str) -> Tuple[List[dict], int, str]:
    """
    Parse Lenovo DCSC Quote-like sheet and also return the sheet name actually used.

    Returns (rows, header_row_index, chosen_sheet_name) with header_row_index 0-based.
    Each row dict: Part number, Product Description, Qty, Price, Export Control,
    __row_index__ (1-based Excel row).
    data_only=False because Total Part Price column has formulas we must not
    collapse to cached values unintentionally.

    Use this from the adapter when downstream code (e.g. annotated_writer) needs
    to re-open the same sheet. Use parse_excel() when only rows + header index matter.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=False)
    try:
        sheetnames = wb.sheetnames
        chosen: Optional[str] = None
        all_rows: Optional[List[Tuple]] = None
        header_row_index: int = 0
        col_map: Dict[str, int] = {}

        for name in ordered_sheet_candidates(sheetnames):
            ws = wb[name]
            rows = list(ws.iter_rows(values_only=True))
            header_hit = find_lenovo_header_in_rows(rows[: _header_scan_limit()])
            if header_hit is None:
                continue
            header_row_index, _ = header_hit
            header_tuple = rows[header_row_index]
            col_map = {}
            for idx, val in enumerate(header_tuple):
                if val is not None:
                    col_map[str(val).strip()] = idx
            missing = [m for m in HEADER_MARKERS if m not in col_map]
            if missing:
                raise ValueError(
                    f"Lenovo header row missing columns {missing} in sheet {name!r} "
                    f"of {filepath}. Found headers: {sorted(col_map.keys())}"
                )
            chosen = name
            all_rows = rows
            break

        if chosen is None or all_rows is None:
            raise ValueError(
                f"No Lenovo DCSC header row found (need all markers {HEADER_MARKERS}) "
                f"in any sheet of {filepath}. Sheets: {sheetnames}"
            )

        pn_idx = col_map["Part number"]
        data_start = len(all_rows)
        for i in range(header_row_index + 1, len(all_rows)):
            if _part_number_nonempty(all_rows[i], pn_idx):
                data_start = i
                break

        rows_out: List[dict] = []
        seen_data = False

        for i in range(data_start, len(all_rows)):
            raw = all_rows[i]
            excel_row_1based = i + 1

            if _first_nonempty_stripped(raw).lower().startswith(_STOP_SENTINEL):
                break

            if not seen_data:
                if not _part_number_nonempty(raw, pn_idx):
                    continue
                seen_data = True

            rows_out.append(
                {
                    "Part number": _cell_at(raw, col_map["Part number"]),
                    "Product Description": _cell_at(raw, col_map["Product Description"]),
                    "Qty": _cell_at(raw, col_map["Qty"]),
                    "Price": _cell_at(raw, col_map["Price"]),
                    "Export Control": _cell_at(raw, col_map["Export Control"]),
                    "__row_index__": excel_row_1based,
                }
            )

        return (rows_out, header_row_index, chosen)
    finally:
        wb.close()


def parse_excel(filepath: str) -> Tuple[List[dict], int]:
    """
    Parse Lenovo DCSC Quote-like sheet.

    Returns (rows, header_row_index) with header_row_index 0-based.
    Thin wrapper over parse_excel_with_sheet() that drops the chosen sheet name --
    keeps the historical 2-tuple contract for callers that don't need it.
    """
    rows, header_row_index, _ = parse_excel_with_sheet(filepath)
    return (rows, header_row_index)
