"""
HPE QuoteBuilder BOM Excel parser.
Sheet 'BOM', header row 0 (no preamble).
Columns determined by name via col_map (not by index — columns are unstable across files).
"""

import openpyxl
from pathlib import Path
from typing import List, Tuple


def parse_excel(filepath: str) -> Tuple[List[dict], int]:
    """
    Parse HPE QuoteBuilder BOM Excel file.

    Returns (rows, header_row_index).
    rows: list[dict] — each dict contains row fields + '__row_index__' (1-based Excel row number).
    header_row_index: always 0 (HPE BOM has no preamble).

    Column detection: by header name with .strip() (not by index).
    End-of-data detection: first non-empty cell == "total" (case-insensitive).
    hp4.xlsx has no Total row — parser correctly finishes by exhausting sheet rows.
    """
    path = Path(filepath)
    if not path.exists():
        raise FileNotFoundError(f"Excel file not found: {filepath}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if "BOM" not in wb.sheetnames:
            raise ValueError(
                f"Sheet 'BOM' not found in {filepath}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb["BOM"]
        rows_iter = ws.iter_rows(values_only=True)

        # Row 0: header (header_row_index = 0)
        header_row = next(rows_iter, None)
        if header_row is None:
            return ([], 0)

        # Build col_map by header name (strip whitespace)
        col_map = {}
        for idx, val in enumerate(header_row):
            if val is not None:
                col_map[str(val).strip()] = idx

        rows: List[dict] = []
        excel_row_number = 2  # header is row 1 (1-based), data starts at row 2

        for raw_row in rows_iter:
            # End-of-data detection by first non-empty cell (columns unstable — not row[0])
            first_nonempty = next((str(v).strip() for v in raw_row if v is not None), "")
            if first_nonempty.lower() == "total":
                break

            # Skip completely empty rows
            if not any(v is not None for v in raw_row):
                excel_row_number += 1
                continue

            row_dict = {}
            for col_name, idx in col_map.items():
                val = raw_row[idx] if idx < len(raw_row) else None
                row_dict[col_name] = val
            row_dict["__row_index__"] = excel_row_number
            rows.append(row_dict)

            excel_row_number += 1

        return (rows, 0)
    finally:
        wb.close()
