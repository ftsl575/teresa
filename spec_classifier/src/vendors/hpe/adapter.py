import openpyxl
from src.vendors.base import VendorAdapter
from src.vendors.hpe.parser import parse_excel
from src.vendors.hpe.normalizer import normalize_hpe_rows


class HPEAdapter(VendorAdapter):
    def __init__(self, config: dict = None):
        self._config = config or {}

    def can_parse(self, path: str) -> bool:
        """
        Positive signature only:
        1. Sheet "BOM" present in workbook.
        2. Row 1 contains both "Product #" and "Product Description".
        """
        wb = openpyxl.load_workbook(path, read_only=True)
        try:
            if "BOM" not in wb.sheetnames:
                return False
            ws = wb["BOM"]
            row1 = next(ws.iter_rows(max_row=1, values_only=True), ())
            cells = {str(v).strip() for v in row1 if v is not None}
            return "Product #" in cells and "Product Description" in cells
        finally:
            wb.close()

    def parse(self, filepath: str):
        return parse_excel(filepath)

    def normalize(self, raw_rows):
        return normalize_hpe_rows(raw_rows)

    def get_rules_file(self):
        vendor_rules = self._config.get("vendor_rules", {})
        return vendor_rules.get("hpe", "rules/hpe_rules.yaml")

    def get_vendor_stats(self, normalized_rows: list) -> dict:
        factory_integrated_count = sum(
            1 for r in normalized_rows if getattr(r, "is_factory_integrated", False)
        )
        config_names = {
            getattr(r, "config_name", "")
            for r in normalized_rows
            if getattr(r, "config_name", "")
        }
        return {
            "factory_integrated_count": factory_integrated_count,
            "config_groups_count": len(config_names),
        }

    def generates_branded_spec(self) -> bool:
        return False
