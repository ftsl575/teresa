"""
Branded specification document: groups items by server (BASE) and entity type sections.
Output mirrors YADRO-style format with company brand colors.
"""

from pathlib import Path
from typing import List

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.core.normalizer import NormalizedRow, RowKind
from src.core.classifier import ClassificationResult, EntityType
from src.core.state_detector import State


# ── Toggle flags ────────────────────────────────────────────────────────────
SHOW_ABSENT_BLOCK = True   # Set False to hide "Не установлено" section

# ── Brand colors ────────────────────────────────────────────────────────────
C_PRIMARY       = "2A90D2"   # Primary blue — title, table header, BASE row
C_PRIMARY_LIGHT = "5DAFF8"   # Reserved
C_WHITE         = "FFFFFF"
C_ZEBRA         = "F0F7FF"   # Light blue tint for alternating rows

SECTION_COLORS = {
    EntityType.HW:       "B7DEFF",
    EntityType.CONFIG:   "E6E6FE",
    EntityType.SOFTWARE: "C7B7FE",
    EntityType.SERVICE:  "B3F4EC",
    EntityType.LOGISTIC: "E8E8E8",
    EntityType.NOTE:     "F5F5F5",
    EntityType.UNKNOWN:  "FFD0D0",
}

SECTION_LABELS = {
    EntityType.HW:       "Оборудование",
    EntityType.CONFIG:   "Конфигурация",
    EntityType.SOFTWARE: "Программное обеспечение",
    EntityType.SERVICE:  "Сервис и гарантия",
    EntityType.LOGISTIC: "Логистика",
    EntityType.NOTE:     "Примечания",
    EntityType.UNKNOWN:  "Нераспознанные позиции",
}

ABSENT_SECTION_COLOR = "F5E6E6"
ABSENT_SECTION_LABEL = "Не установлено / Не выбрано"

SECTION_ORDER = [
    EntityType.HW,
    EntityType.CONFIG,
    EntityType.SOFTWARE,
    EntityType.SERVICE,
    EntityType.LOGISTIC,
    EntityType.NOTE,
    EntityType.UNKNOWN,
]

COL_A_WIDTH  = 3.5
COL_B_WIDTH  = 23.0
COL_C_WIDTH  = 70.0
COL_D_WIDTH  = 12.0


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, size=11, color="000000", italic=False) -> Font:
    return Font(name="Arial", bold=bold, size=size, color=color, italic=italic)


def _thin_border() -> Border:
    side = Side(style="thin", color="D0D0D0")
    return Border(bottom=side)


def _write_title(ws, row: int, source_name: str) -> int:
    cell = ws.cell(row=row, column=3, value=f"Спецификация на основе {source_name}")
    cell.font = Font(name="Arial", bold=True, size=18, color=C_PRIMARY)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 36
    return row + 2


def _write_table_header(ws, row: int) -> int:
    headers = {2: "Артикул", 3: "Наименование", 4: "Кол-во"}
    for col, text in headers.items():
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _fill(C_PRIMARY)
        cell.font = _font(bold=True, size=11, color=C_WHITE)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = _thin_border()
    ws.row_dimensions[row].height = 20
    return row + 1


def _write_base_row(ws, row: int, option_name: str, skus: List[str], qty) -> int:
    ws.cell(row=row, column=2, value=", ".join(skus) if skus else "").fill = _fill(C_PRIMARY)
    ws.cell(row=row, column=3, value=option_name).fill = _fill(C_PRIMARY)
    ws.cell(row=row, column=4, value=qty).fill = _fill(C_PRIMARY)
    for col in (2, 3, 4):
        c = ws.cell(row=row, column=col)
        c.font = _font(bold=True, size=11, color=C_WHITE)
        c.alignment = Alignment(
            horizontal="left" if col < 4 else "center",
            vertical="center",
            wrap_text=False,
        )
    ws.row_dimensions[row].height = 18
    return row + 1


def _write_section_header(ws, row: int, entity_type: EntityType) -> int:
    label = SECTION_LABELS.get(entity_type, entity_type.value)
    bg = SECTION_COLORS.get(entity_type, "E0E0E0")
    for col in (2, 3, 4):
        c = ws.cell(row=row, column=col, value=label if col == 3 else None)
        c.fill = _fill(bg)
        c.font = _font(bold=True, size=10, italic=True)
        c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 16
    return row + 1


def _write_data_row(ws, row: int, skus: List[str], option_name: str, qty, zebra: bool) -> int:
    bg = C_ZEBRA if zebra else C_WHITE
    ws.cell(row=row, column=2, value=", ".join(skus) if skus else "").fill = _fill(bg)
    ws.cell(row=row, column=3, value=option_name).fill = _fill(bg)
    ws.cell(row=row, column=4, value=qty).fill = _fill(bg)
    for col in (2, 3, 4):
        c = ws.cell(row=row, column=col)
        c.font = _font(size=10)
        c.alignment = Alignment(
            horizontal="left" if col < 4 else "center",
            vertical="top",
            wrap_text=True,
        )
    ws.row_dimensions[row].height = None  # Excel autofit on open
    return row + 1


def generate_branded_spec(
    normalized_rows: List[NormalizedRow],
    classification_results: List[ClassificationResult],
    source_filename: str,
    output_path: Path,
) -> Path:
    """
    Build branded specification Excel.

    Groups rows by BASE; within each block renders sections in SECTION_ORDER.
    PRESENT rows only (NOTE/UNKNOWN always included — no state filter).
    ABSENT rows collected into separate block at end of each server block.
    Row height = None (Excel autofit on open).
    """
    output_path = Path(output_path)

    blocks = []
    current = None

    for nrow, result in zip(normalized_rows, classification_results):
        if result.row_kind != RowKind.ITEM:
            continue
        if result.entity_type == EntityType.BASE:
            current = {
                "base": (nrow, result),
                "sections": {et: [] for et in SECTION_ORDER},
                "absent": [],
            }
            blocks.append(current)
            continue
        if current is None:
            continue
        if result.entity_type is None:
            continue

        if result.state == State.ABSENT:
            current["absent"].append((nrow, result))
            continue

        if result.entity_type not in (EntityType.NOTE, EntityType.UNKNOWN):
            if result.state != State.PRESENT:
                continue

        et = result.entity_type
        if et in current["sections"]:
            current["sections"][et].append((nrow, result))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Спецификация"

    ws.column_dimensions["A"].width = COL_A_WIDTH
    ws.column_dimensions["B"].width = COL_B_WIDTH
    ws.column_dimensions["C"].width = COL_C_WIDTH
    ws.column_dimensions["D"].width = COL_D_WIDTH

    stem = Path(source_filename).stem
    current_row = 1
    current_row = _write_title(ws, current_row, stem)
    header_row = current_row
    current_row = _write_table_header(ws, current_row)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    for block in blocks:
        base_nrow, _ = block["base"]
        current_row = _write_base_row(
            ws, current_row,
            option_name=base_nrow.option_name,
            skus=list(base_nrow.skus),
            qty=base_nrow.qty,
        )

        for entity_type in SECTION_ORDER:
            section_rows = block["sections"][entity_type]
            if not section_rows:
                continue
            current_row = _write_section_header(ws, current_row, entity_type)
            for zebra_idx, (nrow, _) in enumerate(section_rows):
                current_row = _write_data_row(
                    ws, current_row,
                    skus=list(nrow.skus),
                    option_name=nrow.option_name,
                    qty=nrow.qty,
                    zebra=(zebra_idx % 2 == 1),
                )

        if SHOW_ABSENT_BLOCK and block["absent"]:
            for col in (2, 3, 4):
                c = ws.cell(row=current_row, column=col,
                            value=ABSENT_SECTION_LABEL if col == 3 else None)
                c.fill = _fill(ABSENT_SECTION_COLOR)
                c.font = _font(bold=True, size=10, italic=True)
                c.alignment = Alignment(horizontal="left", vertical="center")
            ws.row_dimensions[current_row].height = 16
            current_row += 1
            for zebra_idx, (nrow, _) in enumerate(block["absent"]):
                current_row = _write_data_row(
                    ws, current_row,
                    skus=list(nrow.skus),
                    option_name=nrow.option_name,
                    qty=nrow.qty,
                    zebra=(zebra_idx % 2 == 1),
                )

        ws.row_dimensions[current_row].height = 8
        current_row += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
